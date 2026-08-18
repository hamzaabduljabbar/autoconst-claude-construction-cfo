"""Job Cost report (Phase 2, hardened) — deterministic, per project x cost code.

Hardening pass (independent review):
  * `as_of` reporting cutoff is now actually applied to transactions, progress
    updates, and approved budget versions.
  * Unclassified project cost (line has a project but no cost code) is shown as an
    UNCLASSIFIED row per project, with a classification-coverage figure — never
    silently dropped.
  * Recognized actual excludes known non-cost account types (income / asset /
    liability / equity), guarding against a mis-coded balance-sheet line.
  * The unsafe "remaining commitment = committed - all actuals" figure is removed
    (it wrongly subtracted wages/materials unrelated to the subcontract); only the
    safe Committed figure is shown until transactions can be linked to commitments.
  * A Report Info manifest sheet and a Transaction Detail drill-down sheet provide
    visible lineage: import-run ids, source period, reconciliation, workbook hash,
    reporting cutoff, calc-policy version, classification coverage.

All money is integer minor units; all arithmetic is exact; no LLM touches a number.

    build_job_cost_report(db, out_xlsx, as_of=None)
"""

from __future__ import annotations

import datetime as _dt
import sqlite3
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from money import from_minor, fmt_minor

CALC_POLICY_VERSION = "jobcost-0.2.0"
FAC_METHOD = "actual_plus_remaining_budget"

# variance-alert thresholds (explicit, never "sample size sufficient")
ALERT_MIN_BUDGET_MINOR = 1000_00
ALERT_MIN_PCT_BP = 1000            # >=10% complete
ALERT_MIN_ABS_VAR_MINOR = 2000_00  # >=$2,000
ALERT_OVERRUN_RATIO_BP = 500       # overrun >5% OF APPROVED BUDGET

# account types that are NOT cost (excluded from recognized actual)
NONCOST_TYPES = ("income", "revenue", "asset", "liability", "equity",
                 "bank", "accounts receivable", "accounts payable")

_AUD_FMT = '#,##0.00'


def _now() -> str:
    return _dt.datetime.now().isoformat(timespec="seconds")


def _date_clause(as_of, col):
    """Return (sql_fragment, params) applying an optional cutoff on `col`."""
    if as_of:
        return f" AND {col} <= ?", [as_of]
    return "", []


def _approved_budget(conn, pid, ccid, as_of):
    # Highest approved version effective at the cutoff. The v0 baseline (blank
    # approved_date) is always effective; a dated revision counts only once its
    # approval date is on/before the cutoff. With no cutoff, all approved count.
    if as_of:
        cond = "AND (bv.approved_date IS NULL OR bv.approved_date='' OR bv.approved_date <= ?)"
        params = [pid, ccid, as_of]
    else:
        cond = ""
        params = [pid, ccid]
    row = conn.execute(f"""
        SELECT bl.amount_minor FROM budget_lines bl
        JOIN budget_versions bv ON bv.id = bl.budget_version_id
        WHERE bv.project_id=? AND bl.cost_code_id=? AND bv.status='approved' {cond}
        ORDER BY bv.version_no DESC LIMIT 1""", params).fetchone()
    return row[0] if row else 0


def _original_budget(conn, pid, ccid):
    row = conn.execute("""
        SELECT bl.amount_minor FROM budget_lines bl
        JOIN budget_versions bv ON bv.id = bl.budget_version_id
        WHERE bv.project_id=? AND bl.cost_code_id=? AND bv.version_no=0 LIMIT 1""",
        (pid, ccid)).fetchone()
    return row[0] if row else 0


def _latest_pct_bp(conn, pid, ccid, as_of):
    dc, dp = _date_clause(as_of, "pu.as_of_date")
    row = conn.execute(f"""
        SELECT pu.pct_complete, pu.as_of_date, a.progress_method
        FROM progress_updates pu JOIN activities a ON a.id = pu.activity_id
        WHERE a.project_id=? AND a.cost_code_id=? {dc}
        ORDER BY pu.as_of_date DESC LIMIT 1""", [pid, ccid] + dp).fetchone()
    return row if row else (None, None, None)


def _committed(conn, pid, ccid):
    orig = conn.execute("SELECT COALESCE(SUM(original_value_minor),0) FROM commitments "
                        "WHERE project_id=? AND cost_code_id=?", (pid, ccid)).fetchone()[0]
    chg = conn.execute("""SELECT COALESCE(SUM(cc.value_minor),0) FROM commitment_changes cc
        JOIN commitments c ON c.id=cc.commitment_id
        WHERE c.project_id=? AND c.cost_code_id=? AND cc.status='approved'""",
        (pid, ccid)).fetchone()[0]
    return orig + chg


def _actual(conn, pid, ccid, as_of):
    """Recognized actual: posted cost lines, credits net, tax/void excluded,
    non-cost account types excluded, honoring the reporting cutoff."""
    dc, dp = _date_clause(as_of, "h.txn_date")
    noncost = ",".join("?" * len(NONCOST_TYPES))
    code_clause = "l.cost_code_id=?" if ccid is not None else "l.cost_code_id IS NULL"
    params = [pid] + ([ccid] if ccid is not None else []) + list(NONCOST_TYPES) + dp
    return conn.execute(f"""
        SELECT COALESCE(SUM(l.amount_minor),0)
        FROM transaction_lines l
        JOIN transaction_headers h ON h.id = l.header_id
        LEFT JOIN accounts a ON a.id = l.account_id
        WHERE l.project_id=? AND {code_clause} AND l.is_tax=0 AND h.is_void=0
          AND (a.account_type IS NULL OR LOWER(a.account_type) NOT IN ({noncost}))
          {dc}""", params).fetchone()[0]


def compute_rows(conn, as_of=None):
    projects = conn.execute("SELECT id, code, name, status FROM projects ORDER BY code").fetchall()
    result = []
    for pid, code, name, status in projects:
        codes = conn.execute("""
            SELECT DISTINCT cc.id, cc.code, cc.name FROM cost_codes cc WHERE cc.id IN (
                SELECT bl.cost_code_id FROM budget_lines bl
                    JOIN budget_versions bv ON bv.id=bl.budget_version_id WHERE bv.project_id=?
                UNION SELECT l.cost_code_id FROM transaction_lines l
                    WHERE l.project_id=? AND l.cost_code_id IS NOT NULL
                UNION SELECT cost_code_id FROM commitments WHERE project_id=? AND cost_code_id IS NOT NULL
            ) ORDER BY cc.code""", (pid, pid, pid)).fetchall()
        prows = []
        for ccid, cccode, ccname in codes:
            orig = _original_budget(conn, pid, ccid)
            appr = _approved_budget(conn, pid, ccid, as_of)
            actual = _actual(conn, pid, ccid, as_of)
            committed = _committed(conn, pid, ccid)
            pct_bp, pdate, method = _latest_pct_bp(conn, pid, ccid, as_of)
            earned = appr * (pct_bp or 0) // 10000
            variance = earned - actual
            fac = actual + max(appr - earned, 0)
            overrun = actual - earned
            alert = (appr >= ALERT_MIN_BUDGET_MINOR and (pct_bp or 0) >= ALERT_MIN_PCT_BP
                     and overrun >= ALERT_MIN_ABS_VAR_MINOR
                     and appr > 0 and (overrun * 10000) // appr >= ALERT_OVERRUN_RATIO_BP)
            prows.append(dict(cost_code=cccode, description=ccname, original=orig,
                              approved=appr, committed=committed, actual=actual,
                              pct_bp=pct_bp, pdate=pdate, method=method, earned=earned,
                              variance=variance, fac=fac, alert=alert))
        unclassified = _actual(conn, pid, None, as_of)  # project set, no cost code
        classified = sum(r["actual"] for r in prows)
        total_cost = classified + unclassified
        coverage_bp = (classified * 10000 // total_cost) if total_cost else 10000
        result.append(dict(pid=pid, code=code, name=name, status=status, rows=prows,
                           unclassified=unclassified, coverage_bp=coverage_bp))
    return result


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #

_HDRS = ["Project", "Cost Code", "Description", "Original Budget", "Approved Budget",
         "Committed", "Recognized Actual", "% Complete", "As Of", "Progress Method",
         "Earned Value", "Cost Variance", "FAC Method", "Forecast @ Completion", "Alert"]


def _san(v):
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
        return "'" + v
    return v


def build_job_cost_report(db: str, out_xlsx: str, as_of: str | None = None) -> dict:
    conn = sqlite3.connect(db)
    conn.execute("PRAGMA foreign_keys = ON")
    meta = dict(conn.execute("SELECT key,value FROM meta").fetchall())
    currency = meta.get("functional_currency", "AUD")

    runs = conn.execute("SELECT id, adapter, adapter_version, source_period, status "
                        "FROM import_runs ORDER BY id").fetchall()
    last_status = runs[-1][4] if runs else None
    reconciled = last_status == "reconciled"
    blockers = conn.execute(
        "SELECT COUNT(*) FROM data_quality_issues WHERE severity='blocking'").fetchone()[0]

    data = compute_rows(conn, as_of=as_of)

    # portfolio coverage
    tot_class = sum(r["actual"] for p in data for r in p["rows"])
    tot_unclass = sum(p["unclassified"] for p in data)
    port_cost = tot_class + tot_unclass
    port_cov_bp = (tot_class * 10000 // port_cost) if port_cost else 10000

    # Trust requires ALL of: the latest import reconciled, no blocking issues,
    # every project cost classified, AND a full-period view (a date-cutoff view is
    # a subset that was never independently reconciled to a control total at that
    # date, so it is analytical/DRAFT by definition).
    trusted = reconciled and blockers == 0 and tot_unclass == 0 and as_of is None

    run_id = conn.execute(
        "INSERT INTO report_runs(tenant_id, generated_at, trusted, engine_version, notes) "
        "VALUES (?,?,?,?,?)",
        (meta.get("tenant_id"), _now(), 1 if trusted else 0, CALC_POLICY_VERSION,
         f"job cost as_of={as_of or 'latest'}")).lastrowid
    conn.commit()

    wb = Workbook()
    ws = wb.active
    ws.title = "Job Cost"

    bold = Font(bold=True); white = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    proj_fill = PatternFill("solid", fgColor="D9E1F2")
    alert_fill = PatternFill("solid", fgColor="F8CBAD")
    total_fill = PatternFill("solid", fgColor="E2EFDA")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    right = Alignment(horizontal="right")

    title = (f"Job Cost Report — {meta.get('tenant_id')}" if trusted
             else "DRAFT — NOT TRUSTED (see Report Info sheet)")
    ws.append([title])
    ws["A1"].font = Font(bold=True, color="000000" if trusted else "C00000", size=13)
    ws.append([f"As of {as_of or 'latest'} · currency {currency} · "
               f"classification coverage {port_cov_bp/100:.2f}%"])
    ws.append([])
    hr = ws.max_row + 1
    ws.append(_HDRS)
    for c in ws[hr]:
        c.font = white; c.fill = hdr_fill

    def money_cell(r, col, minor):
        cell = ws.cell(row=r, column=col, value=float(from_minor(minor)))
        cell.number_format = _AUD_FMT; cell.alignment = right

    grand = dict(original=0, approved=0, committed=0, actual=0, earned=0, variance=0, fac=0)
    for proj in data:
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=_san(f"{proj['code']} — {proj['name']} ({proj['status']})")).font = bold
        for c in ws[r]:
            c.fill = proj_fill
        psum = dict(original=0, approved=0, committed=0, actual=0, earned=0, variance=0, fac=0)
        for row in proj["rows"]:
            r = ws.max_row + 1
            ws.cell(row=r, column=2, value=_san(row["cost_code"]))
            ws.cell(row=r, column=3, value=_san(row["description"]))
            money_cell(r, 4, row["original"]); money_cell(r, 5, row["approved"])
            money_cell(r, 6, row["committed"]); money_cell(r, 7, row["actual"])
            ws.cell(row=r, column=8, value="" if row["pct_bp"] is None else f'{row["pct_bp"]/100:.0f}%').alignment = right
            ws.cell(row=r, column=9, value=_san(row["pdate"] or ""))
            ws.cell(row=r, column=10, value=_san(row["method"] or ""))
            money_cell(r, 11, row["earned"]); money_cell(r, 12, row["variance"])
            ws.cell(row=r, column=13, value=FAC_METHOD)
            money_cell(r, 14, row["fac"])
            ws.cell(row=r, column=15, value="OVER-BUDGET" if row["alert"] else "")
            if row["alert"]:
                for c in ws[r]:
                    c.fill = alert_fill
            for k in psum:
                psum[k] += row[k]
        # UNCLASSIFIED row (never silently dropped)
        if proj["unclassified"]:
            r = ws.max_row + 1
            ws.cell(row=r, column=2, value="—")
            ws.cell(row=r, column=3, value="UNCLASSIFIED (project cost, no cost code)").font = bold
            money_cell(r, 7, proj["unclassified"])
            ws.cell(row=r, column=15, value=f"coverage {proj['coverage_bp']/100:.1f}%").alignment = right
            for c in ws[r]:
                c.fill = warn_fill
            psum["actual"] += proj["unclassified"]
        # project subtotal
        r = ws.max_row + 1
        ws.cell(row=r, column=3, value=f"Total {proj['code']}").font = bold
        for col, k in ((4, "original"), (5, "approved"), (6, "committed"), (7, "actual"),
                       (11, "earned"), (12, "variance"), (14, "fac")):
            money_cell(r, col, psum[k])
        for c in ws[r]:
            c.fill = total_fill
        for k in grand:
            grand[k] += psum[k]

    r = ws.max_row + 2
    ws.cell(row=r, column=3, value="GRAND TOTAL").font = Font(bold=True, size=12)
    for col, k in ((4, "original"), (5, "approved"), (6, "committed"), (7, "actual"),
                   (11, "earned"), (12, "variance"), (14, "fac")):
        money_cell(r, col, grand[k])

    widths = [34, 11, 34, 15, 15, 14, 16, 11, 12, 20, 15, 15, 26, 20, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = f"A{hr+1}"

    _write_report_info(wb, meta, as_of, trusted, reconciled, blockers, runs,
                       port_cov_bp, tot_class, tot_unclass, run_id, currency)
    _write_transaction_detail(wb, conn, as_of, currency)

    wb.save(out_xlsx)
    conn.close()

    summary = dict(trusted=trusted, reconciled=reconciled, projects=len(data),
                   grand=grand, coverage_bp=port_cov_bp, unclassified=tot_unclass,
                   alerts=sum(1 for p in data for row in p["rows"] if row["alert"]))
    print(f"[jobcost] {'TRUSTED' if trusted else 'DRAFT/NOT-TRUSTED'} · as_of={as_of or 'latest'} · "
          f"{summary['projects']} projects · {summary['alerts']} alerts · "
          f"coverage {port_cov_bp/100:.2f}%")
    print(f"[jobcost] actual {fmt_minor(grand['actual'], currency)} "
          f"(incl unclassified {fmt_minor(tot_unclass, currency)}) vs approved "
          f"{fmt_minor(grand['approved'], currency)}")
    print(f"[jobcost] wrote {out_xlsx}")
    return summary


def _write_report_info(wb, meta, as_of, trusted, reconciled, blockers, runs,
                       cov_bp, classified, unclassified, run_id, currency):
    ws = wb.create_sheet("Report Info")
    ws.append(["Report manifest / lineage"])
    ws["A1"].font = Font(bold=True, size=13)
    rows = [
        ("Tenant", meta.get("tenant_id")),
        ("Generated at", _now()),
        ("Reporting cutoff (as_of)", as_of or "latest (all data)"),
        ("Report run id", run_id),
        ("Calc policy version", CALC_POLICY_VERSION),
        ("Functional currency", currency),
        ("TRUSTED", "YES" if trusted else "NO"),
        ("  reconciled import", "yes" if reconciled else "no"),
        ("  blocking data-quality issues", blockers),
        ("  classification coverage", f"{cov_bp/100:.1f}%"),
        ("  classified cost", fmt_minor(classified, currency)),
        ("  unclassified cost", fmt_minor(unclassified, currency)),
        ("Workbook name", meta.get("workbook_name", "—")),
        ("Workbook hash (sha256)", meta.get("workbook_hash", "—")),
    ]
    for k, v in rows:
        ws.append([k, _san(str(v))])
    ws.append([])
    ws.append(["Import runs contributing to this report"])
    ws.append(["id", "adapter", "adapter_version", "source_period", "status"])
    for rid, adapter, ver, period, status in runs:
        ws.append([rid, adapter, ver, period or "—", status])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 40
    for col in "CDE":
        ws.column_dimensions[col].width = 18


def _write_transaction_detail(wb, conn, as_of, currency):
    """Drill-down: every cost line behind the numbers, with a stable key."""
    ws = wb.create_sheet("Transaction Detail")
    ws.append(["Transaction detail (cost lines feeding this report)"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append(["Line Key", "Date", "Project", "Cost Code", "Account", "Vendor",
               "Amount", "Memo", "Source Txn"])
    for c in ws[2]:
        c.font = Font(bold=True)
    dc, dp = _date_clause(as_of, "h.txn_date")
    noncost = ",".join("?" * len(NONCOST_TYPES))
    q = f"""
        SELECT l.id, h.txn_date, p.code, cc.code, a.name, pt.name,
               l.amount_minor, l.memo, h.native_transaction_id
        FROM transaction_lines l
        JOIN transaction_headers h ON h.id = l.header_id
        LEFT JOIN projects p ON p.id = l.project_id
        LEFT JOIN cost_codes cc ON cc.id = l.cost_code_id
        LEFT JOIN accounts a ON a.id = l.account_id
        LEFT JOIN parties pt ON pt.id = h.party_id
        WHERE l.project_id IS NOT NULL AND l.is_tax=0 AND h.is_void=0
          AND (a.account_type IS NULL OR LOWER(a.account_type) NOT IN ({noncost}))
          {dc}
        ORDER BY p.code, cc.code, h.txn_date"""
    for lid, date, pcode, cccode, acc, vendor, amt, memo, src in conn.execute(q, list(NONCOST_TYPES) + dp):
        key = f"{src}:{lid}"
        cell = [_san(key), _san(date or ""), _san(pcode or ""), _san(cccode or "UNCLASSIFIED"),
                _san(acc or ""), _san(vendor or ""), float(from_minor(amt)),
                _san((memo or "")[:80]), _san(src or "")]
        ws.append(cell)
        ws.cell(row=ws.max_row, column=7).number_format = _AUD_FMT
    for col, w in zip("ABCDEFGHI", (20, 12, 10, 14, 18, 22, 14, 50, 14)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
