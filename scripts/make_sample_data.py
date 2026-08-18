"""Generate a realistic, internally-consistent sample dataset for Construction CFO.

Emits into sample-data/:
    canonical-transactions.csv   the documented canonical transaction-LINE format
    qbo-transaction-detail.csv   a QuickBooks-Online-flavoured export of the same data
    control-totals.csv           per-account and per-project totals (reconciliation target)
    project-input.xlsx           the structured project-and-forecast input workbook

The control totals are COMPUTED from the same lines that are written out, so an
importer that reconciles to them is checking real arithmetic, not a hand-typed
number. The scenario deliberately exercises the hard cases:
    * a single bill split across two projects + a company-level (no-project) line
    * a GST/tax line kept separate from cost
    * a credit note (negative amount)
    * a voided transaction (present, flagged, excluded from totals)
    * a vague-memo line the classifier should abstain on
"""

from __future__ import annotations

import csv
import datetime as _dt
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "sample-data"
OUT.mkdir(exist_ok=True)

TENANT = "Northwind Construction Pty Ltd"
PLATFORM = "qbo"
CURRENCY = "AUD"

# --- reference data ---------------------------------------------------------

PROJECTS = [
    # native_id, code, name, status
    ("PRJ-01", "NGW", "Northgate Warehouse", "active"),
    ("PRJ-02", "BAC", "Bayside Aged Care", "active"),
    ("PRJ-03", "CVB", "Clearview Bridge", "active"),
]

ACCOUNTS = [
    ("ACC-MAT", "Materials", "cogs"),
    ("ACC-SUB", "Subcontractors", "cogs"),
    ("ACC-PLT", "Plant Hire", "cogs"),
    ("ACC-WAG", "Wages", "cogs"),
    ("ACC-OHD", "Office Overhead", "expense"),
    ("ACC-GST", "GST Payable", "liability"),
]

COST_CODES = [
    ("03-30-00", "Cast-in-place concrete"),
    ("31-23-00", "Earthworks"),
    ("05-12-00", "Structural steel"),
    ("07-40-00", "Roofing & cladding"),
    ("32-13-00", "Concrete paving"),
    ("01-50-00", "Temporary facilities"),
]

VENDORS = [
    "Boral Concrete", "CivilWorks Earthmoving", "SteelFab Australia",
    "Metro Roofing Co", "Coates Hire", "OfficeWorks", "Northwind Payroll",
]

# --- transactions -----------------------------------------------------------
# Each entry is a header plus one or more lines.
# amount is a string decimal (never a float). project=None -> company-level.

def L(account, project, cost_code, amount, memo, tax=False):
    return dict(account=account, project=project, cost_code=cost_code,
                amount=amount, memo=memo, tax=tax)

TXNS = [
    dict(ttype="Bill", num="B-1001", vendor="Boral Concrete", date="2026-07-03",
         void=False, lines=[
            L("Materials", "PRJ-01", "03-30-00", "12400.00", "N32 concrete supply — slab pour A"),
            L("Materials", "PRJ-01", "03-30-00", "3100.00", "N32 concrete supply — slab pour B"),
            L("GST Payable", None, None, "1550.00", "GST on B-1001", tax=True),
         ]),
    # One bill split across TWO projects + a company-level delivery line
    dict(ttype="Bill", num="B-1002", vendor="SteelFab Australia", date="2026-07-05",
         void=False, lines=[
            L("Materials", "PRJ-01", "05-12-00", "8600.00", "310UB structural steel — portal frames"),
            L("Materials", "PRJ-02", "05-12-00", "5400.00", "SHS columns — canopy"),
            L("Materials", None, None, "220.00", "Freight — split delivery"),
            L("GST Payable", None, None, "1422.00", "GST on B-1002", tax=True),
         ]),
    dict(ttype="Bill", num="B-1003", vendor="CivilWorks Earthmoving", date="2026-07-08",
         void=False, lines=[
            L("Subcontractors", "PRJ-01", "31-23-00", "18750.00", "Bulk earthworks — progress claim 2"),
            L("GST Payable", None, None, "1875.00", "GST on B-1003", tax=True),
         ]),
    dict(ttype="Bill", num="B-1004", vendor="Coates Hire", date="2026-07-10",
         void=False, lines=[
            L("Plant Hire", "PRJ-01", "31-23-00", "3200.00", "Excavator hire — 2 weeks"),
            L("Plant Hire", "PRJ-03", "31-23-00", "2800.00", "Crane hire — girder lift"),
            L("GST Payable", None, None, "600.00", "GST on B-1004", tax=True),
         ]),
    dict(ttype="Bill", num="B-1005", vendor="Metro Roofing Co", date="2026-07-14",
         void=False, lines=[
            L("Subcontractors", "PRJ-02", "07-40-00", "22300.00", "Colorbond cladding — east elevation"),
            L("GST Payable", None, None, "2230.00", "GST on B-1005", tax=True),
         ]),
    # A vague memo — classifier should ABSTAIN (UNCLASSIFIED), not force a code
    dict(ttype="Bill", num="B-1006", vendor="OfficeWorks", date="2026-07-15",
         void=False, lines=[
            L("Materials", "PRJ-02", None, "640.00", "Sundries"),
            L("GST Payable", None, None, "64.00", "GST on B-1006", tax=True),
         ]),
    dict(ttype="Bill", num="B-1007", vendor="CivilWorks Earthmoving", date="2026-07-18",
         void=False, lines=[
            L("Subcontractors", "PRJ-03", "32-13-00", "14200.00", "Approach paving — south abutment"),
            L("GST Payable", None, None, "1420.00", "GST on B-1007", tax=True),
         ]),
    # A credit note — negative (roofing over-claim corrected)
    dict(ttype="CreditNote", num="CN-2001", vendor="Metro Roofing Co", date="2026-07-22",
         void=False, lines=[
            L("Subcontractors", "PRJ-02", "07-40-00", "-1800.00", "Credit — over-claim adjustment PC5"),
            L("GST Payable", None, None, "-180.00", "GST on CN-2001", tax=True),
         ]),
    # Payroll journal (no tax line)
    dict(ttype="JournalEntry", num="JE-3001", vendor="Northwind Payroll", date="2026-07-25",
         void=False, lines=[
            L("Wages", "PRJ-01", "03-30-00", "6800.00", "Site labour — concrete crew wk30"),
            L("Wages", "PRJ-02", "07-40-00", "4200.00", "Site labour — fixing crew wk30"),
            L("Wages", None, None, "3500.00", "Head office salaries wk30"),
         ]),
    dict(ttype="Bill", num="B-1008", vendor="OfficeWorks", date="2026-07-28",
         void=False, lines=[
            L("Office Overhead", None, None, "900.00", "Monthly software + stationery"),
            L("GST Payable", None, None, "90.00", "GST on B-1008", tax=True),
         ]),
    # A VOIDED bill — present in the file, flagged, excluded from control totals
    dict(ttype="Bill", num="B-1009", vendor="Coates Hire", date="2026-07-29",
         void=True, lines=[
            L("Plant Hire", "PRJ-03", "31-23-00", "1500.00", "Scissor lift — VOIDED duplicate"),
            L("GST Payable", None, None, "150.00", "GST on B-1009", tax=True),
         ]),
]


# --- writers ----------------------------------------------------------------

def write_canonical():
    """Clean, documented canonical transaction-LINE format."""
    path = OUT / "canonical-transactions.csv"
    fields = [
        "tenant", "platform", "txn_type", "txn_id", "line_id", "date",
        "vendor", "account", "project", "cost_code", "tax_code",
        "amount", "currency", "is_tax", "is_void", "memo",
    ]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for t in TXNS:
            for i, ln in enumerate(t["lines"], start=1):
                w.writerow({
                    "tenant": TENANT, "platform": PLATFORM,
                    "txn_type": t["ttype"], "txn_id": t["num"],
                    "line_id": f'{t["num"]}-L{i}', "date": t["date"],
                    "vendor": t["vendor"], "account": ln["account"],
                    "project": ln["project"] or "", "cost_code": ln["cost_code"] or "",
                    "tax_code": "GST" if ln["tax"] else "",
                    "amount": ln["amount"], "currency": CURRENCY,
                    "is_tax": 1 if ln["tax"] else 0,
                    "is_void": 1 if t["void"] else 0,
                    "memo": ln["memo"],
                })
    return path


def write_qbo():
    """A QuickBooks-Online-flavoured export: MM/DD/YYYY dates, $-formatted amounts."""
    path = OUT / "qbo-transaction-detail.csv"
    fields = ["Date", "Transaction Type", "No.", "Name", "Customer/Project",
              "Memo/Description", "Account", "Amount"]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for t in TXNS:
            proj_name = {p[0]: p[2] for p in PROJECTS}
            for ln in t["lines"]:
                d = _dt.date.fromisoformat(t["date"]).strftime("%m/%d/%Y")
                amt = Decimal(ln["amount"])
                amt_s = f"-${abs(amt):,.2f}" if amt < 0 else f"${amt:,.2f}"
                w.writerow({
                    "Date": d, "Transaction Type": t["ttype"], "No.": t["num"],
                    "Name": t["vendor"],
                    "Customer/Project": proj_name.get(ln["project"], "") if ln["project"] else "",
                    "Memo/Description": ("VOID: " if t["void"] else "") + ln["memo"],
                    "Account": ln["account"], "Amount": amt_s,
                })
    return path


def write_qbo_messy():
    """A realistic 'Transaction Detail by Account' export, with the junk a real
    QBO CSV carries: title rows, grouped-by-account (account is a GROUP HEADER,
    not a per-row column), 'Total for <account>' subtotals, blank separator rows,
    a running Balance column, $-formatted amounts, MM/DD/YYYY dates, and voided
    transactions shown as $0.00. The adapter must survive all of it.
    """
    path = OUT / "qbo-transaction-detail-messy.csv"
    proj_name = {p[0]: p[2] for p in PROJECTS}

    def money(a: Decimal) -> str:
        return f"-${abs(a):,.2f}" if a < 0 else f"${a:,.2f}"

    # bucket lines by account, preserving transaction identity
    groups: dict[str, list] = defaultdict(list)
    for t in TXNS:
        for ln in t["lines"]:
            groups[ln["account"]].append((t, ln))

    rows: list[list[str]] = []
    rows.append([TENANT, "", "", "", "", "", "", ""])
    rows.append(["Transaction Detail by Account", "", "", "", "", "", "", ""])
    rows.append(["July 1-31, 2026", "", "", "", "", "", "", ""])
    rows.append(["", "", "", "", "", "", "", ""])
    # column header row (note: NO 'Account' column — account is the group header)
    rows.append(["Date", "Transaction Type", "Num", "Name",
                 "Customer/Project", "Memo/Description", "Amount", "Balance"])

    # keep account order stable and familiar
    account_order = [a[1] for a in ACCOUNTS]
    for acc in account_order:
        if acc not in groups:
            continue
        rows.append([acc, "", "", "", "", "", "", ""])  # group header
        balance = Decimal(0)
        subtotal = Decimal(0)
        for t, ln in groups[acc]:
            amt = Decimal(ln["amount"])
            shown = Decimal("0.00") if t["void"] else amt
            balance += shown
            subtotal += shown
            d = _dt.date.fromisoformat(t["date"]).strftime("%m/%d/%Y")
            memo = ("Voided: " if t["void"] else "") + ln["memo"]
            rows.append([
                d, t["ttype"], t["num"], t["vendor"],
                proj_name.get(ln["project"], "") if ln["project"] else "",
                memo, money(shown), money(balance),
            ])
        rows.append(["", "", "", "", "", f"Total for {acc}", money(subtotal), ""])
        rows.append(["", "", "", "", "", "", "", ""])  # blank separator

    with path.open("w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)
    return path


def write_control_totals():
    """Per-account and per-project totals, COMPUTED from the non-void lines.

    This is the reconciliation target. Tax lines are reported separately so a
    reconciler can prove tax was never folded into cost.
    """
    path = OUT / "control-totals.csv"
    by_account = defaultdict(Decimal)
    by_project = defaultdict(Decimal)
    tax_total = Decimal(0)
    grand_cost = Decimal(0)

    for t in TXNS:
        if t["void"]:
            continue
        for ln in t["lines"]:
            amt = Decimal(ln["amount"])
            if ln["tax"]:
                tax_total += amt
                continue
            by_account[ln["account"]] += amt
            key = ln["project"] or "(unassigned)"
            by_project[key] += amt
            grand_cost += amt

    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["dimension", "key", "amount", "currency"])
        for acc, v in sorted(by_account.items()):
            w.writerow(["account", acc, f"{v:.2f}", CURRENCY])
        for prj, v in sorted(by_project.items()):
            w.writerow(["project", prj, f"{v:.2f}", CURRENCY])
        w.writerow(["tax", "GST Payable", f"{tax_total:.2f}", CURRENCY])
        w.writerow(["total", "cost_ex_tax", f"{grand_cost:.2f}", CURRENCY])
    return path, grand_cost, tax_total


def write_input_workbook():
    """The structured project-and-forecast input workbook the contractor maintains."""
    path = OUT / "project-input.xlsx"
    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")

    def sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for c in ws[1]:
            c.font = hdr_font
            c.fill = hdr_fill
        for r in rows:
            ws.append(r)
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[chr(64 + i)].width = max(14, len(str(h)) + 2)
        return ws

    wb.remove(wb.active)  # drop default sheet

    # Meta / schema version
    sheet("_Meta", ["key", "value"], [
        ["workbook_schema_version", "0.1.0"],
        ["tenant", TENANT],
        ["functional_currency", CURRENCY],
        ["prepared_by", "Sample generator"],
    ])

    sheet("Projects", ["native_id", "code", "name", "status"],
          [[p[0], p[1], p[2], p[3]] for p in PROJECTS])

    sheet("CostCodes", ["code", "name"], [[c[0], c[1]] for c in COST_CODES])

    # Activities: project x cost_code with quantities, dates, progress method
    sheet("Activities",
          ["project_code", "cost_code", "name", "quantity", "unit",
           "planned_start", "planned_finish", "progress_method"],
          [
           ["NGW", "31-23-00", "Bulk earthworks", 4500, "m3", "2026-07-01", "2026-07-20", "physical_qty"],
           ["NGW", "03-30-00", "Slab & footings", 320, "m3", "2026-07-10", "2026-08-05", "physical_qty"],
           ["NGW", "05-12-00", "Structural steel", 42, "t", "2026-07-25", "2026-08-20", "physical_qty"],
           ["BAC", "05-12-00", "Canopy steel", 8, "t", "2026-07-01", "2026-07-15", "physical_qty"],
           ["BAC", "07-40-00", "Cladding", 1200, "m2", "2026-07-10", "2026-08-10", "physical_qty"],
           ["CVB", "31-23-00", "Abutment earthworks", 900, "m3", "2026-07-05", "2026-07-25", "physical_qty"],
           ["CVB", "32-13-00", "Approach paving", 650, "m2", "2026-07-15", "2026-08-05", "physical_qty"],
          ])

    # Progress updates (drives earned value)
    sheet("Progress",
          ["project_code", "cost_code", "as_of_date", "pct_complete", "recorded_by"],
          [
           ["NGW", "31-23-00", "2026-07-31", 90, "Site PM"],
           ["NGW", "03-30-00", "2026-07-31", 35, "Site PM"],
           ["NGW", "05-12-00", "2026-07-31", 10, "Site PM"],
           ["BAC", "05-12-00", "2026-07-31", 80, "Site PM"],
           ["BAC", "07-40-00", "2026-07-31", 45, "Site PM"],
           ["CVB", "31-23-00", "2026-07-31", 60, "Site PM"],
           ["CVB", "32-13-00", "2026-07-31", 25, "Site PM"],
          ])

    # Budgets: version 0 (original) budget lines per project x cost_code
    sheet("Budgets",
          ["project_code", "budget_version", "status", "cost_code", "amount", "label"],
          [
           ["NGW", 0, "approved", "31-23-00", 20000, "Original tender"],
           ["NGW", 0, "approved", "03-30-00", 42000, "Original tender"],
           ["NGW", 0, "approved", "05-12-00", 60000, "Original tender"],
           ["BAC", 0, "approved", "05-12-00", 12000, "Original tender"],
           ["BAC", 0, "approved", "07-40-00", 38000, "Original tender"],
           ["CVB", 0, "approved", "31-23-00", 9000, "Original tender"],
           ["CVB", 0, "approved", "32-13-00", 26000, "Original tender"],
           # An APPROVED revision to one line (extra scope) — baseline preserved above
           ["BAC", 1, "approved", "07-40-00", 44000, "Approved revision 1 — added west elevation"],
          ])

    # Head contracts
    sheet("Contracts",
          ["project_code", "client", "contract_value", "payment_terms_days",
           "claim_frequency", "retention_pct", "retention_release"],
          [
           ["NGW", "Northgate Developments", 480000, 30, "monthly", 5.0, "practical_completion"],
           ["BAC", "Bayside Care Group", 610000, 45, "monthly", 5.0, "defects_liability_end"],
           ["CVB", "State Roads Authority", 390000, 30, "monthly", 10.0, "practical_completion"],
          ])

    sheet("ContractChanges",
          ["project_code", "seq_no", "description", "value", "status", "approved_date"],
          [
           ["BAC", 1, "Variation — west elevation cladding", 34000, "approved", "2026-07-12"],
           ["NGW", 1, "Variation — extra roller door", 8500, "pending", ""],
          ])

    # Subcontracts / commitments
    sheet("Commitments",
          ["project_code", "subcontractor", "cost_code", "original_value",
           "payment_terms_days", "retention_pct"],
          [
           ["NGW", "CivilWorks Earthmoving", "31-23-00", 19000, 30, 5.0],
           ["BAC", "Metro Roofing Co", "07-40-00", 40000, 30, 5.0],
           ["CVB", "CivilWorks Earthmoving", "32-13-00", 25000, 30, 5.0],
          ])

    sheet("CommitmentChanges",
          ["project_code", "subcontractor", "seq_no", "description", "value", "status"],
          [
           ["BAC", "Metro Roofing Co", 1, "West elevation add", 6000, "approved"],
          ])

    # Business-level cash inputs (blocking inputs for the cash-flow forecast)
    sheet("CashInputs",
          ["key", "value", "note"],
          [
           ["opening_bank_balance", 85000, "as at forecast start"],
           ["forecast_start_date", "2026-08-01", ""],
           ["payroll_weekly", 14500, "site + office"],
           ["overhead_monthly", 22000, "head office run rate"],
           ["gst_next_payment_date", "2026-08-28", ""],
           ["gst_next_payment_amount", 18000, "estimated BAS"],
          ])

    wb.save(path)
    return path


def main():
    p1 = write_canonical()
    p2 = write_qbo()
    p5 = write_qbo_messy()
    p3, grand, tax = write_control_totals()
    p4 = write_input_workbook()
    n_lines = sum(len(t["lines"]) for t in TXNS)
    n_void = sum(1 for t in TXNS if t["void"])
    print(f"[sample] {len(TXNS)} transactions, {n_lines} lines "
          f"({n_void} voided txn excluded from totals)")
    print(f"[sample] cost ex-tax total: {grand:.2f} {CURRENCY} | GST: {tax:.2f} {CURRENCY}")
    for p in (p1, p2, p5, p3, p4):
        print(f"[sample] wrote {p.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
