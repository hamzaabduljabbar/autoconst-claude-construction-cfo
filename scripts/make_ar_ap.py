"""Generate A/R and A/P aging exports for Summit Ridge (money in / money out).

These stand in for the QuickBooks 'Customer Balance Detail' (open receivables)
and 'Vendor Balance Detail' (open payables) reports — the data that turns the
cash-flow report from a schedule-based ESTIMATE into an invoice-based forecast.

Deterministic; tells a coherent story:
  * open client claims due across the next quarter (APR the late payer is overdue)
  * open supplier/subcontractor bills due over the next 8 weeks
Written as clean CSVs (the messy-grouped parser is already proven on the txn export).
"""

from __future__ import annotations

import csv
import datetime as _dt
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SR = ROOT / "integration-data" / "summit-ridge"
START = _dt.date(2026, 8, 1)


def _d(offset_days):
    return (START + _dt.timedelta(days=offset_days)).isoformat()


def main():
    wb = load_workbook(SR / "project-input.xlsx", data_only=True)
    contracts = {r[0]: dict(client=r[1], value=r[2], terms=r[3])
                 for r in wb["Contracts"].iter_rows(min_row=2, values_only=True)}

    # --- A/R: open client claims (money owed TO us) ---
    # (project, doc_no, issue_offset, due_offset, amount) — negative offset = before forecast start
    ar_plan = {
        "WHX": [("PC-14", -20, 10, 96000)],
        "MCF": [("PC-11", -12, 18, 88000)],
        "APR": [("PC-09", -55, -25, 132000), ("PC-10", -20, 10, 74000)],  # late payer: one overdue
        "ISL": [("PC-08", -8, 22, 61000)],
        "MBR": [("PC-05", 5, 35, 145000), ("PC-06", 33, 63, 120000)],     # active, ongoing claims
        "SCH": [("PC-04", 12, 42, 132000)],
    }
    ar_rows = []
    for pcode, items in ar_plan.items():
        c = contracts.get(pcode, {})
        for doc_no, iss, due, amt in items:
            ar_rows.append([c.get("client", pcode), pcode, "Claim", f"{pcode}-{doc_no}",
                            _d(iss), _d(due), f"{amt}.00", f"{amt}.00"])

    # --- A/P: open supplier / subcontractor bills (money WE owe) ---
    ap_plan = {
        "MBR": [("Redgum Civil", "B-7001", -10, 20, 88000),
                ("Coates Hire", "B-7002", 2, 32, 26500)],
        "SCH": [("Delta Electrical", "B-7010", -5, 25, 54000),
                ("Boral Concrete", "B-7011", 8, 38, 41000)],
        "APR": [("Precision Carpentry", "B-7020", -30, 0, 33000)],        # due now
        "ISL": [("SteelFab Australia", "B-7030", -2, 28, 47000)],
        "WHX": [("PrimeCo Painting", "B-7040", 6, 36, 18500)],
        "MCF": [("Axis Interiors", "B-7050", 15, 45, 62000)],
    }
    ap_rows = []
    for pcode, items in ap_plan.items():
        for vendor, doc_no, iss, due, amt in items:
            ap_rows.append([vendor, pcode, "Bill", doc_no, _d(iss), _d(due),
                            f"{amt}.00", f"{amt}.00"])

    ar_path = SR / "qbo-exports" / "ar-aging.csv"
    ap_path = SR / "qbo-exports" / "ap-aging.csv"
    with ar_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["customer", "project", "doc_type", "doc_no", "issue_date", "due_date", "amount", "open_balance"])
        w.writerows(ar_rows)
    with ap_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["vendor", "project", "doc_type", "doc_no", "issue_date", "due_date", "amount", "open_balance"])
        w.writerows(ap_rows)

    ar_tot = sum(Decimal(r[7]) for r in ar_rows)
    ap_tot = sum(Decimal(r[7]) for r in ap_rows)
    print(f"[ar/ap] {len(ar_rows)} open receivables = {ar_tot:,.2f} · {len(ap_rows)} open payables = {ap_tot:,.2f}")
    print(f"[ar/ap] wrote {ar_path.relative_to(ROOT)} and {ap_path.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
