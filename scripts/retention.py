"""Retention tracker (Phase 4) — deterministic.

Per project × contract:
  * retention HELD BY THE CLIENT from us (an asset — we get it back on release)
  * retention WE WITHHOLD from our subcontractors (a liability — we pay it out)
  * the date each retention tranche becomes claimable
  * a running outstanding-retention position

Retention held by client   = contract retention% × revenue certified to date
Revenue certified to date   = contract value × overall % complete
Overall % complete          = Σ earned value / Σ approved budget  (budget-weighted)
Sub retention withheld       = Σ  sub retention% × cost certified to that sub

Release date is derived from the retention_release trigger:
  practical_completion   -> project planned finish
  defects_liability_end  -> planned finish + defects-liability period (default 365d)

All money is integer minor units; no LLM touches a number.

    build_retention_report(db, out_xlsx, as_of=None, defects_days=365)
"""

from __future__ import annotations

import datetime as _dt
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from money import from_minor, fmt_minor
from jobcost import _approved_budget, _latest_pct_bp, NONCOST_TYPES

_AUD = '#,##0.00'


def _now():
    return _dt.datetime.now().isoformat(timespec="seconds")


def _add_days(iso, days):
    try:
        return (_dt.date.fromisoformat(iso) + _dt.timedelta(days=days)).isoformat()
    except (ValueError, TypeError):
        return ""


def _project_finish(conn, pid):
    row = conn.execute("SELECT MAX(planned_finish) FROM activities WHERE project_id=?", (pid,)).fetchone()
    return row[0] if row and row[0] else ""


def _overall_pct_bp(conn, pid, as_of):
    """Budget-weighted % complete across the project's cost codes (basis points)."""
    codes = conn.execute("""SELECT DISTINCT cost_code_id FROM budget_lines bl
        JOIN budget_versions bv ON bv.id=bl.budget_version_id WHERE bv.project_id=?""", (pid,)).fetchall()
    tot_appr = tot_earned = 0
    for (ccid,) in codes:
        appr = _approved_budget(conn, pid, ccid, as_of)
        pct_bp, _, _ = _latest_pct_bp(conn, pid, ccid, as_of)
        tot_appr += appr
        tot_earned += appr * (pct_bp or 0) // 10000
    return (tot_earned * 10000 // tot_appr) if tot_appr else 0


def _sub_retention(conn, pid, as_of):
    """Σ retention% × recognized cost to each sub. Deduped by VENDOR so a vendor
    with several commitments on one project is not counted more than once."""
    noncost = ",".join("?" * len(NONCOST_TYPES))
    dc, dp = ("", [])
    if as_of:
        dc, dp = " AND h.txn_date <= ?", [as_of]
    # one retention rate per vendor on this project (max of their commitments)
    by_party = {}
    for party_id, rbp in conn.execute(
            "SELECT party_id, retention_pct_bp FROM commitments WHERE project_id=?", (pid,)):
        if not party_id or not rbp:
            continue
        by_party[party_id] = max(by_party.get(party_id, 0), rbp)
    total = 0
    for party_id, rbp in by_party.items():
        cost = conn.execute(f"""SELECT COALESCE(SUM(l.amount_minor),0)
            FROM transaction_lines l JOIN transaction_headers h ON h.id=l.header_id
            LEFT JOIN accounts a ON a.id=l.account_id
            WHERE l.project_id=? AND h.party_id=? AND l.is_tax=0 AND h.is_void=0
              AND (a.account_type IS NULL OR LOWER(a.account_type) NOT IN ({noncost})){dc}""",
            [pid, party_id] + list(NONCOST_TYPES) + dp).fetchone()[0]
        total += cost * rbp // 10000
    return total


def _retention_status(release, today, horizon):
    """Bucket a release date relative to now."""
    if not release:
        return "NO DATE"
    if release < today:
        return "CLAIMABLE NOW (overdue)"
    if release <= horizon:
        return "DUE WITHIN 90 DAYS"
    return "NOT YET DUE"


def compute_retention(conn, as_of=None, defects_days=365):
    today = as_of or _dt.date.today().isoformat()
    horizon = _add_days(today, 90)
    out = []
    for pid, code, name, status in conn.execute(
            "SELECT id, code, name, status FROM projects ORDER BY code"):
        c = conn.execute("""SELECT contract_value_minor, retention_pct_bp, retention_release_trigger, party_id
            FROM contracts WHERE project_id=? LIMIT 1""", (pid,)).fetchone()
        if not c:
            continue
        cval, rbp, trigger, client_id = c
        client = conn.execute("SELECT name FROM parties WHERE id=?", (client_id,)).fetchone()
        client = client[0] if client else ""
        pct_bp = _overall_pct_bp(conn, pid, as_of)
        revenue_certified = (cval or 0) * pct_bp // 10000
        client_retention = revenue_certified * (rbp or 0) // 10000
        finish = _project_finish(conn, pid)
        release = finish if trigger == "practical_completion" else _add_days(finish, defects_days)
        sub_ret = _sub_retention(conn, pid, as_of)
        out.append(dict(code=code, name=name, status=status, client=client,
                        contract=cval or 0, pct_bp=pct_bp, revenue_certified=revenue_certified,
                        client_retention=client_retention, trigger=trigger, release=release,
                        sub_retention=sub_ret, net=client_retention - sub_ret,
                        claim_status=_retention_status(release, today, horizon)))
    return out


def build_retention_report(db: str, out_xlsx: str, as_of=None, defects_days=365) -> dict:
    conn = sqlite3.connect(db)
    conn.execute("PRAGMA foreign_keys = ON")
    meta = dict(conn.execute("SELECT key,value FROM meta").fetchall())
    currency = meta.get("functional_currency", "AUD")
    rows = compute_retention(conn, as_of, defects_days)
    conn.close()

    today = as_of or _dt.date.today().isoformat()
    horizon = _add_days(today, 90)

    wb = Workbook(); ws = wb.active; ws.title = "Retention"
    white = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="2F5496")
    total_fill = PatternFill("solid", fgColor="E2EFDA"); soon_fill = PatternFill("solid", fgColor="FFF2CC")
    overdue_fill = PatternFill("solid", fgColor="F8CBAD")
    right = Alignment(horizontal="right")

    ws.append([f"Estimated Retention Position — {meta.get('tenant_id')}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"As of {today} · {currency} · ESTIMATED from % complete — not from certified "
               f"claims or a retention ledger. Release dates are planned, not certified."])
    ws.append([])
    hdrs = ["Project", "Client", "Contract Value", "% Complete", "Est. Revenue Certified",
            "Est. Client Retention", "Release Trigger", "Planned Release Date",
            "Est. Sub Retention Withheld", "Net Position", "Status"]
    hr = ws.max_row + 1
    ws.append(hdrs)
    for c in ws[hr]:
        c.font = white; c.fill = fill

    def money(r, col, minor):
        cell = ws.cell(row=r, column=col, value=float(from_minor(minor)))
        cell.number_format = _AUD; cell.alignment = right

    tot = dict(contract=0, revenue_certified=0, client_retention=0, sub_retention=0, net=0)
    for row in rows:
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=f"{row['code']} — {row['name']}")
        ws.cell(row=r, column=2, value=row["client"])
        money(r, 3, row["contract"])
        ws.cell(row=r, column=4, value=f"{row['pct_bp']/100:.0f}%").alignment = right
        money(r, 5, row["revenue_certified"])
        money(r, 6, row["client_retention"])
        ws.cell(row=r, column=7, value=row["trigger"])
        ws.cell(row=r, column=8, value=row["release"])
        money(r, 9, row["sub_retention"])
        money(r, 10, row["net"])
        ws.cell(row=r, column=11, value=row["claim_status"])
        if row["claim_status"].startswith("CLAIMABLE NOW"):
            for c in ws[r]:
                c.fill = overdue_fill
        elif row["claim_status"] == "DUE WITHIN 90 DAYS":
            for c in ws[r]:
                c.fill = soon_fill
        for k in tot:
            tot[k] += row[k]

    r = ws.max_row + 1
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    money(r, 3, tot["contract"]); money(r, 5, tot["revenue_certified"])
    money(r, 6, tot["client_retention"]); money(r, 9, tot["sub_retention"]); money(r, 10, tot["net"])
    for c in ws[r]:
        c.fill = total_fill

    for i, w in enumerate([30, 22, 16, 11, 16, 22, 20, 13, 26, 20, 13], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = f"A{hr+1}"
    wb.save(out_xlsx)

    overdue = sum(1 for r in rows if r["claim_status"].startswith("CLAIMABLE NOW"))
    print(f"[retention] {len(rows)} contracts · EST client retention {fmt_minor(tot['client_retention'], currency)} "
          f"· EST sub retention {fmt_minor(tot['sub_retention'], currency)} · {overdue} claimable now (overdue)")
    print(f"[retention] net position {fmt_minor(tot['net'], currency)} · wrote {out_xlsx}")
    return dict(rows=rows, totals=tot)
