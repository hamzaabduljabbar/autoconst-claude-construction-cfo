"""Ingest + reconciliation gate (Phase 1).

    run_ingest(db, source, adapter, projects_workbook) -> import_run_id
    run_reconcile(db, import_run_id, control_csv)      -> (ok: bool, issues)

Guarantees:
  * Idempotent — re-importing the same file inserts no duplicate rows.
  * Update-aware — a changed transaction replaces its prior lines (not appended).
  * Voids/credits kept and flagged, excluded from reconciliation totals.
  * Tax lines never folded into cost totals.
  * Unassigned-project costs stay visible under the '(unassigned)' key.
  * Nothing is trusted until imported totals equal the source control totals.
"""

from __future__ import annotations

import csv
import datetime as _dt
import hashlib
import json
import sqlite3
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

from adapters import ADAPTERS
from money import to_minor, from_minor, fmt_minor

UNASSIGNED = "(unassigned)"


def _now() -> str:
    return _dt.datetime.now().isoformat(timespec="seconds")


def _sig(*parts) -> str:
    return hashlib.sha256("|".join(str(p) for p in parts).encode("utf-8")).hexdigest()


def _file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


# --------------------------------------------------------------------------- #
# reference resolution (create-on-first-sight, cached)
# --------------------------------------------------------------------------- #

class _Refs:
    def __init__(self, conn: sqlite3.Connection, tenant: str, projects_workbook=None):
        self.c = conn
        self.tenant = tenant
        self._acc: dict[str, int] = {}
        self._party: dict[str, int] = {}
        self._cc: dict[str, int] = {}
        self._proj: dict[str, int | None] = {}
        self._proj_alias: dict[str, str] = {}   # name/code -> native_id
        if projects_workbook:
            self._load_project_aliases(projects_workbook)

    def _load_project_aliases(self, workbook_path):
        from openpyxl import load_workbook
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        if "Projects" not in wb.sheetnames:
            return
        ws = wb["Projects"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        header = [str(h).strip().lower() if h else "" for h in rows[0]]
        idx = {h: i for i, h in enumerate(header)}
        for r in rows[1:]:
            if not r or all(v is None for v in r):
                continue
            native = str(r[idx["native_id"]]).strip() if "native_id" in idx and r[idx["native_id"]] else None
            if not native:
                continue
            for key in ("native_id", "code", "name"):
                if key in idx and r[idx[key]]:
                    self._proj_alias[str(r[idx[key]]).strip()] = native

    def account(self, name: str) -> int | None:
        if not name:
            return None
        if name in self._acc:
            return self._acc[name]
        row = self.c.execute(
            "SELECT id FROM accounts WHERE tenant_id=? AND name=?",
            (self.tenant, name)).fetchone()
        if row:
            aid = row[0]
        else:
            aid = self.c.execute(
                "INSERT INTO accounts(tenant_id, native_id, name) VALUES (?,?,?)",
                (self.tenant, name, name)).lastrowid
        self._acc[name] = aid
        return aid

    def party(self, name: str) -> int | None:
        if not name:
            return None
        if name in self._party:
            return self._party[name]
        row = self.c.execute(
            "SELECT id FROM parties WHERE tenant_id=? AND name=? AND role='vendor'",
            (self.tenant, name)).fetchone()
        if row:
            pid = row[0]
        else:
            pid = self.c.execute(
                "INSERT INTO parties(tenant_id, native_id, name, role) VALUES (?,?,?, 'vendor')",
                (self.tenant, name, name)).lastrowid
        self._party[name] = pid
        return pid

    def cost_code(self, code: str) -> int | None:
        if not code:
            return None
        if code in self._cc:
            return self._cc[code]
        row = self.c.execute(
            "SELECT id FROM cost_codes WHERE tenant_id=? AND code=?",
            (self.tenant, code)).fetchone()
        if row:
            cid = row[0]
        else:
            cid = self.c.execute(
                "INSERT INTO cost_codes(tenant_id, code, name) VALUES (?,?,?)",
                (self.tenant, code, code)).lastrowid
        self._cc[code] = cid
        return cid

    def project(self, raw: str) -> tuple[int | None, str]:
        """Return (project_id or None, canonical_native_id_or_UNASSIGNED)."""
        if not raw:
            return None, UNASSIGNED
        native = self._proj_alias.get(raw, raw)  # resolve name/code -> native
        if native in self._proj:
            return self._proj[native], native
        row = self.c.execute(
            "SELECT id FROM projects WHERE tenant_id=? AND native_id=?",
            (self.tenant, native)).fetchone()
        if row:
            pid = row[0]
        else:
            pid = self.c.execute(
                "INSERT INTO projects(tenant_id, native_id, code, name, status) "
                "VALUES (?,?,?,?, 'active')",
                (self.tenant, native, native, native)).lastrowid
        self._proj[native] = pid
        return pid, native


# --------------------------------------------------------------------------- #
# ingest
# --------------------------------------------------------------------------- #

def run_ingest(db: str, source: str, adapter: str, projects_workbook=None,
               source_platform="qbo") -> int:
    if adapter not in ADAPTERS:
        raise ValueError(f"unknown adapter '{adapter}'; have {list(ADAPTERS)}")
    parse, adapter_version = ADAPTERS[adapter]

    conn = sqlite3.connect(db)
    conn.execute("PRAGMA foreign_keys = ON")
    tenant = dict(conn.execute("SELECT key,value FROM meta").fetchall())["tenant_id"]
    currency = dict(conn.execute("SELECT key,value FROM meta").fetchall()).get(
        "functional_currency", "AUD")

    src_path = Path(source)
    # source system
    ss = conn.execute(
        "SELECT id FROM source_systems WHERE tenant_id=? AND platform=?",
        (tenant, source_platform)).fetchone()
    ss_id = ss[0] if ss else conn.execute(
        "INSERT INTO source_systems(tenant_id, platform, company_name, accounting_basis) "
        "VALUES (?,?,?, 'accrual')", (tenant, source_platform, tenant)).lastrowid

    run_id = conn.execute(
        "INSERT INTO import_runs(source_system_id, source_file_name, source_file_hash, "
        "adapter, adapter_version, imported_at, status) VALUES (?,?,?,?,?,?, 'ingesting')",
        (ss_id, src_path.name, _file_hash(src_path), adapter, adapter_version, _now())
    ).lastrowid

    refs = _Refs(conn, tenant, projects_workbook)

    # group normalized lines by business transaction identity
    headers: dict[tuple, dict] = {}
    raw_count = 0
    for nl in parse(source):
        raw_count += 1
        key = (nl["txn_type"], nl["txn_id"])
        h = headers.setdefault(key, dict(
            txn_type=nl["txn_type"], txn_id=nl["txn_id"], date=nl["date"],
            vendor=nl["vendor"], currency=nl["currency"],
            is_void=nl["is_void"], lines=[]))
        h["is_void"] = h["is_void"] or nl["is_void"]
        h["lines"].append(nl)

    kept = updated = inserted = skipped = 0
    fp_dupes = 0

    for (ttype, txn_id), h in headers.items():
        party_id = refs.party(h["vendor"])
        # build resolved line tuples + a content signature for change detection
        seen_fp: dict[str, int] = {}
        resolved = []
        header_total = 0
        for nl in h["lines"]:
            acc_id = refs.account(nl["account"])
            proj_id, proj_native = refs.project(nl["project_raw"])
            cc_id = refs.cost_code(nl["cost_code"])
            amt_minor = to_minor(nl["amount"], scale=2)
            header_total += amt_minor
            native_line_id = nl["native_line_id"]
            fp = None
            if native_line_id is None:
                fp = _sig(nl["account"], proj_native, nl["cost_code"],
                          nl["amount"], nl["memo"], nl["is_tax"])
                if fp in seen_fp:
                    seen_fp[fp] += 1
                    fp = f"{fp}#{seen_fp[fp]}"   # disambiguate identical lines
                    fp_dupes += 1
                else:
                    seen_fp[fp] = 0
            resolved.append(dict(
                native_line_id=native_line_id, line_fingerprint=fp,
                account_id=acc_id, project_id=proj_id, cost_code_id=cc_id,
                tax_code=nl["tax_code"], amount_minor=amt_minor,
                is_tax=1 if nl["is_tax"] else 0, memo=nl["memo"]))

        content_sig = _sig(h["date"], h["vendor"], h["is_void"], header_total,
                           tuple(sorted((r["native_line_id"] or r["line_fingerprint"],
                                         r["amount_minor"]) for r in resolved)))

        existing = conn.execute(
            "SELECT id FROM transaction_headers WHERE tenant_id=? AND source_platform=? "
            "AND native_transaction_type=? AND native_transaction_id=?",
            (tenant, source_platform, ttype, txn_id)).fetchone()

        if existing:
            hid = existing[0]
            prior = conn.execute(
                "SELECT content_sig FROM transaction_headers WHERE id=?", (hid,)).fetchone()
            if prior and prior[0] == content_sig:
                skipped += 1
                continue  # idempotent: nothing changed
            # changed -> replace lines and refresh header
            conn.execute("DELETE FROM transaction_lines WHERE header_id=?", (hid,))
            conn.execute(
                "UPDATE transaction_headers SET import_run_id=?, party_id=?, txn_date=?, "
                "currency=?, header_total_minor=?, is_void=?, content_sig=? WHERE id=?",
                (run_id, party_id, h["date"], h["currency"], header_total,
                 1 if h["is_void"] else 0, content_sig, hid))
            updated += 1
        else:
            hid = conn.execute(
                "INSERT INTO transaction_headers(import_run_id, tenant_id, source_platform, "
                "native_transaction_type, native_transaction_id, party_id, txn_date, currency, "
                "header_total_minor, is_void, content_sig) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (run_id, tenant, source_platform, ttype, txn_id, party_id, h["date"],
                 h["currency"], header_total, 1 if h["is_void"] else 0, content_sig)).lastrowid
            inserted += 1

        for r in resolved:
            conn.execute(
                "INSERT INTO transaction_lines(header_id, native_line_id, line_fingerprint, "
                "account_id, project_id, cost_code_id, tax_code, amount_minor, is_tax, memo) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (hid, r["native_line_id"], r["line_fingerprint"], r["account_id"],
                 r["project_id"], r["cost_code_id"], r["tax_code"], r["amount_minor"],
                 r["is_tax"], r["memo"]))
        kept += len(resolved)

    if fp_dupes:
        conn.execute(
            "INSERT INTO data_quality_issues(import_run_id, severity, category, message, created_at) "
            "VALUES (?, 'info', 'quarantine', ?, ?)",
            (run_id, f"{fp_dupes} line(s) with no native id shared a fingerprint and were "
                     f"disambiguated by occurrence", _now()))

    conn.execute(
        "UPDATE import_runs SET row_count_raw=?, row_count_kept=?, status='ingested' WHERE id=?",
        (raw_count, kept, run_id))
    conn.commit()

    print(f"[ingest] {src_path.name} via {adapter} ({adapter_version})")
    print(f"[ingest] {len(headers)} transactions -> inserted={inserted} "
          f"updated={updated} skipped(idempotent)={skipped}; {kept} lines kept")
    conn.close()
    return run_id


# --------------------------------------------------------------------------- #
# reconciliation gate
# --------------------------------------------------------------------------- #

def _load_controls(control_csv: str):
    acc, prj = {}, {}
    tax = total = None
    with open(control_csv, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            dim, key, amt = row["dimension"], row["key"], to_minor(row["amount"])
            if dim == "account":
                acc[key] = amt
            elif dim == "project":
                prj[key] = amt
            elif dim == "tax":
                tax = amt
            elif dim == "total":
                total = amt
    return acc, prj, tax, total


def run_reconcile(db: str, control_csv: str, report=True):
    conn = sqlite3.connect(db)
    conn.execute("PRAGMA foreign_keys = ON")
    currency = dict(conn.execute("SELECT key,value FROM meta").fetchall()).get(
        "functional_currency", "AUD")

    # totals from NON-void, NON-tax lines
    acc_rows = conn.execute("""
        SELECT a.name, SUM(l.amount_minor)
        FROM transaction_lines l
        JOIN transaction_headers h ON h.id = l.header_id
        LEFT JOIN accounts a ON a.id = l.account_id
        WHERE h.is_void = 0 AND l.is_tax = 0
        GROUP BY a.name""").fetchall()
    prj_rows = conn.execute("""
        SELECT COALESCE(p.native_id, ?) AS pk, SUM(l.amount_minor)
        FROM transaction_lines l
        JOIN transaction_headers h ON h.id = l.header_id
        LEFT JOIN projects p ON p.id = l.project_id
        WHERE h.is_void = 0 AND l.is_tax = 0
        GROUP BY pk""", (UNASSIGNED,)).fetchall()
    tax_row = conn.execute("""
        SELECT COALESCE(SUM(l.amount_minor), 0)
        FROM transaction_lines l JOIN transaction_headers h ON h.id = l.header_id
        WHERE h.is_void = 0 AND l.is_tax = 1""").fetchone()

    got_acc = {name: amt for name, amt in acc_rows}
    got_prj = {pk: amt for pk, amt in prj_rows}
    got_tax = tax_row[0]

    ctl_acc, ctl_prj, ctl_tax, ctl_total = _load_controls(control_csv)

    issues = []
    def cmp(dim, key, got, want):
        if got != want:
            issues.append(dict(severity="blocking", category="reconciliation",
                               message=f"{dim} '{key}': imported {fmt_minor(got, currency)} "
                                       f"!= control {fmt_minor(want, currency)}"))

    for key, want in ctl_acc.items():
        cmp("account", key, got_acc.get(key, 0), want)
    for key, want in ctl_prj.items():
        cmp("project", key, got_prj.get(key, 0), want)
    if ctl_tax is not None:
        cmp("tax", "GST", got_tax, ctl_tax)
    if ctl_total is not None:
        cmp("total", "cost_ex_tax", sum(got_acc.values()), ctl_total)

    # cross-foot: account grand total must equal project grand total
    if sum(got_acc.values()) != sum(got_prj.values()):
        issues.append(dict(severity="blocking", category="reconciliation",
                           message="account grand total != project grand total (cross-foot failed)"))

    ok = not issues
    now = _now()
    run_id = conn.execute("SELECT MAX(id) FROM import_runs").fetchone()[0]
    for iss in issues:
        conn.execute(
            "INSERT INTO data_quality_issues(import_run_id, severity, category, message, created_at) "
            "VALUES (?,?,?,?,?)", (run_id, iss["severity"], iss["category"], iss["message"], now))
    conn.execute("UPDATE import_runs SET status=? WHERE id=?",
                 ("reconciled" if ok else "unreconciled", run_id))
    conn.commit()

    if report:
        print(f"[reconcile] account totals: {len(ctl_acc)} checked | "
              f"project totals: {len(ctl_prj)} checked")
        print(f"[reconcile] cost ex-tax imported: "
              f"{fmt_minor(sum(got_acc.values()), currency)} | "
              f"GST: {fmt_minor(got_tax, currency)}")
        if ok:
            print("[reconcile] RECONCILED — reports may be trusted")
        else:
            print(f"[reconcile] UNRECONCILED — {len(issues)} discrepancy(ies):")
            for iss in issues:
                print(f"           - {iss['message']}")
    conn.close()
    return ok, issues
