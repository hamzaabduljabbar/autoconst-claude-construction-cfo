"""Load the structured project-and-forecast input workbook into the schema.

Populates the project-side tables the job-cost report (and later phases) read:
cost_codes, projects (code/name/status), activities, progress_updates,
budget_versions + budget_lines, contracts + contract_changes, commitments.

    load_workbook_into_db(db, workbook_path)

Money cells arrive from openpyxl as floats; they are stringified before
conversion so no binary-float value ever reaches a stored amount.
"""

from __future__ import annotations

import hashlib
import sqlite3
from decimal import Decimal

from openpyxl import load_workbook

from ingest import _Refs
from money import to_minor


def _clear_workbook_scope(conn, tenant: str):
    """Idempotency: remove the workbook-derived rows for this tenant's projects
    before reloading, so re-running load-workbook never doubles records. Reference
    tables (accounts, cost_codes, projects, activities) are keyed and upserted, so
    they are left in place; the accumulating tables are cleared."""
    pids = [r[0] for r in conn.execute(
        "SELECT id FROM projects WHERE tenant_id=?", (tenant,)).fetchall()]
    if not pids:
        return
    q = ",".join("?" * len(pids))
    conn.execute(f"DELETE FROM progress_updates WHERE activity_id IN "
                 f"(SELECT id FROM activities WHERE project_id IN ({q}))", pids)
    conn.execute(f"DELETE FROM commitment_changes WHERE commitment_id IN "
                 f"(SELECT id FROM commitments WHERE project_id IN ({q}))", pids)
    conn.execute(f"DELETE FROM commitments WHERE project_id IN ({q})", pids)
    conn.execute(f"DELETE FROM contract_changes WHERE contract_id IN "
                 f"(SELECT id FROM contracts WHERE project_id IN ({q}))", pids)
    conn.execute(f"DELETE FROM contracts WHERE project_id IN ({q})", pids)
    conn.execute(f"DELETE FROM budget_lines WHERE budget_version_id IN "
                 f"(SELECT id FROM budget_versions WHERE project_id IN ({q}))", pids)
    conn.execute(f"DELETE FROM budget_versions WHERE project_id IN ({q})", pids)


def _money(v) -> int:
    if v is None or v == "":
        return 0
    return to_minor(Decimal(str(v)), scale=2)


def _rows(ws):
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return []
    hdr = [str(h).strip() if h is not None else "" for h in data[0]]
    out = []
    for r in data[1:]:
        if r is None or all(c is None for c in r):
            continue
        out.append({hdr[i]: r[i] for i in range(len(hdr))})
    return out


def load_workbook_into_db(db: str, workbook_path: str) -> dict:
    conn = sqlite3.connect(db)
    conn.execute("PRAGMA foreign_keys = ON")
    tenant = dict(conn.execute("SELECT key,value FROM meta").fetchall())["tenant_id"]
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    refs = _Refs(conn, tenant, workbook_path)

    # idempotency: clear prior workbook-derived rows before reloading
    _clear_workbook_scope(conn, tenant)

    # record workbook identity for the report manifest / lineage
    with open(workbook_path, "rb") as f:
        wb_hash = hashlib.sha256(f.read()).hexdigest()
    conn.executemany("INSERT OR REPLACE INTO meta(key,value) VALUES (?,?)",
                     [("workbook_hash", wb_hash), ("workbook_name", str(workbook_path))])

    def client(name):
        """Resolve a head-contract client as a CUSTOMER party (not a vendor)."""
        if not name:
            return None
        row = conn.execute(
            "SELECT id FROM parties WHERE tenant_id=? AND name=? AND role='customer'",
            (tenant, name)).fetchone()
        if row:
            return row[0]
        return conn.execute(
            "INSERT INTO parties(tenant_id, native_id, name, role) VALUES (?,?,?, 'customer')",
            (tenant, name, name)).lastrowid

    counts = {}

    # --- cost codes (names) ---
    if "CostCodes" in wb.sheetnames:
        n = 0
        for r in _rows(wb["CostCodes"]):
            code = str(r["code"]).strip()
            cid = refs.cost_code(code)
            conn.execute("UPDATE cost_codes SET name=? WHERE id=?", (r.get("name") or code, cid))
            n += 1
        counts["cost_codes"] = n

    # --- projects (code / name / status from the master sheet) ---
    if "Projects" in wb.sheetnames:
        n = 0
        for r in _rows(wb["Projects"]):
            native = str(r["native_id"]).strip()
            pid, _ = refs.project(native)
            conn.execute("UPDATE projects SET code=?, name=?, status=? WHERE id=?",
                         (r.get("code"), r.get("name"), r.get("status"), pid))
            n += 1
        counts["projects"] = n

    # --- activities (+ progress method) ---
    act_id: dict[tuple, int] = {}
    if "Activities" in wb.sheetnames:
        n = 0
        for r in _rows(wb["Activities"]):
            pid, _ = refs.project(str(r["project_code"]).strip())
            cid = refs.cost_code(str(r["cost_code"]).strip())
            row = conn.execute(
                "SELECT id FROM activities WHERE project_id=? AND cost_code_id=?", (pid, cid)).fetchone()
            if row:
                aid = row[0]
            else:
                aid = conn.execute(
                    "INSERT INTO activities(project_id, cost_code_id, name, unit, "
                    "planned_start, planned_finish, progress_method) VALUES (?,?,?,?,?,?,?)",
                    (pid, cid, r.get("name"), r.get("unit"),
                     str(r.get("planned_start") or ""), str(r.get("planned_finish") or ""),
                     r.get("progress_method"))).lastrowid
            act_id[(pid, cid)] = aid
            n += 1
        counts["activities"] = n

    # --- progress updates (percent -> basis points, 100% = 10000) ---
    if "Progress" in wb.sheetnames:
        n = 0
        for r in _rows(wb["Progress"]):
            pid, _ = refs.project(str(r["project_code"]).strip())
            cid = refs.cost_code(str(r["cost_code"]).strip())
            aid = act_id.get((pid, cid))
            if aid is None:
                aid = conn.execute(
                    "INSERT INTO activities(project_id, cost_code_id) VALUES (?,?)",
                    (pid, cid)).lastrowid
                act_id[(pid, cid)] = aid
            pct_bp = int(round(float(r.get("pct_complete") or 0) * 100))
            conn.execute(
                "INSERT INTO progress_updates(activity_id, as_of_date, pct_complete, recorded_by) "
                "VALUES (?,?,?,?)",
                (aid, str(r.get("as_of_date") or ""), pct_bp, r.get("recorded_by")))
            n += 1
        counts["progress_updates"] = n

    # --- budgets (versioned) ---
    if "Budgets" in wb.sheetnames:
        bv_id: dict[tuple, int] = {}
        n = 0
        for r in _rows(wb["Budgets"]):
            pid, _ = refs.project(str(r["project_code"]).strip())
            ver = int(r["budget_version"])
            key = (pid, ver)
            if key not in bv_id:
                row = conn.execute(
                    "SELECT id FROM budget_versions WHERE project_id=? AND version_no=?", (pid, ver)).fetchone()
                bv_id[key] = row[0] if row else conn.execute(
                    "INSERT INTO budget_versions(project_id, version_no, label, status) VALUES (?,?,?,?)",
                    (pid, ver, r.get("label"), r.get("status"))).lastrowid
            cid = refs.cost_code(str(r["cost_code"]).strip())
            conn.execute(
                "INSERT OR REPLACE INTO budget_lines(budget_version_id, cost_code_id, amount_minor, note) "
                "VALUES (?,?,?,?)", (bv_id[key], cid, _money(r["amount"]), r.get("label")))
            n += 1
        counts["budget_lines"] = n

    # --- contracts + changes ---
    if "Contracts" in wb.sheetnames:
        contract_id: dict[int, int] = {}
        n = 0
        for r in _rows(wb["Contracts"]):
            pid, _ = refs.project(str(r["project_code"]).strip())
            party = client(r.get("client"))
            cid = conn.execute(
                "INSERT INTO contracts(project_id, party_id, contract_value_minor, payment_terms_days, "
                "claim_frequency, retention_pct_bp, retention_release_trigger, median_days_late) "
                "VALUES (?,?,?,?,?,?,?,?)",
                (pid, party, _money(r["contract_value"]), r.get("payment_terms_days"),
                 r.get("claim_frequency"), int(round(float(r.get("retention_pct") or 0) * 100)),
                 r.get("retention_release"), int(r.get("median_days_late") or 0))).lastrowid
            contract_id[pid] = cid
            n += 1
        counts["contracts"] = n

        if "ContractChanges" in wb.sheetnames:
            m = 0
            for r in _rows(wb["ContractChanges"]):
                pid, _ = refs.project(str(r["project_code"]).strip())
                cid = contract_id.get(pid)
                if cid is None:
                    continue
                conn.execute(
                    "INSERT INTO contract_changes(contract_id, seq_no, description, value_minor, "
                    "status, approved_date) VALUES (?,?,?,?,?,?)",
                    (cid, int(r["seq_no"]), r.get("description"), _money(r["value"]),
                     r.get("status"), str(r.get("approved_date") or ""))); m += 1
            counts["contract_changes"] = m

    # --- commitments ---
    if "Commitments" in wb.sheetnames:
        n = 0
        for r in _rows(wb["Commitments"]):
            pid, _ = refs.project(str(r["project_code"]).strip())
            party = refs.party(r.get("subcontractor")) if r.get("subcontractor") else None
            cc = refs.cost_code(str(r["cost_code"]).strip()) if r.get("cost_code") else None
            conn.execute(
                "INSERT INTO commitments(project_id, party_id, cost_code_id, original_value_minor, "
                "payment_terms_days, retention_pct_bp) VALUES (?,?,?,?,?,?)",
                (pid, party, cc, _money(r["original_value"]), r.get("payment_terms_days"),
                 int(round(float(r.get("retention_pct") or 0) * 100)))); n += 1
        counts["commitments"] = n

    # --- cash-flow inputs (key/value) -> meta under a cash_ prefix ---
    if "CashInputs" in wb.sheetnames:
        n = 0
        for r in _rows(wb["CashInputs"]):
            key = str(r.get("key") or "").strip()
            if not key:
                continue
            conn.execute("INSERT OR REPLACE INTO meta(key,value) VALUES (?,?)",
                         (f"cash_{key}", str(r.get("value"))))
            n += 1
        counts["cash_inputs"] = n

    conn.commit()
    conn.close()
    print(f"[workbook] loaded: " + ", ".join(f"{k}={v}" for k, v in counts.items()))
    return counts
