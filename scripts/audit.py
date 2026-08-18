"""Expense leak audit (Phase 6) — deterministic review alerts, not accusations.

Surfaces three kinds of thing a CFO scans for, from posted cost lines only:

  1. Possible duplicate bills   — same vendor + amount (+ memo) on separate bills
  2. Unusually large costs       — a line far above that vendor's normal, by a
                                    robust median/MAD test (not the mean), with a
                                    minimum sample size so small vendors don't false-fire
  3. Vendor concentration        — where a supply agreement / bulk deal might pay off

Everything is labelled a REVIEW ALERT. The tool never asserts fraud or error — it
points a human at the handful of lines worth eyeballing.

    build_expense_audit(db, out_xlsx) -> summary
"""

from __future__ import annotations

import datetime as _dt
import sqlite3
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from money import from_minor, fmt_minor

NONCOST_TYPES = ("income", "revenue", "asset", "liability", "equity",
                 "bank", "accounts receivable", "accounts payable")

DUP_WINDOW_DAYS = 14
SPIKE_MIN_SAMPLE = 5          # need >=5 lines for a vendor before judging "unusual"
SPIKE_MOD_Z = 3.5            # modified z-score threshold (robust)
CONCENTRATION_MIN = 150_000_00   # flag vendors above AUD 150k total spend
_AUD = '#,##0.00'


def _median(xs):
    s = sorted(xs)
    n = len(s)
    if n == 0:
        return 0
    m = n // 2
    return s[m] if n % 2 else (s[m - 1] + s[m]) / 2


def _cost_lines(conn):
    noncost = ",".join("?" * len(NONCOST_TYPES))
    return conn.execute(f"""
        SELECT l.id, h.native_transaction_id, h.txn_date, pt.name, a.name, p.code,
               l.amount_minor, l.memo
        FROM transaction_lines l
        JOIN transaction_headers h ON h.id = l.header_id
        LEFT JOIN parties pt ON pt.id = h.party_id
        LEFT JOIN accounts a ON a.id = l.account_id
        LEFT JOIN projects p ON p.id = l.project_id
        WHERE l.is_tax=0 AND h.is_void=0 AND l.amount_minor > 0
          AND (a.account_type IS NULL OR LOWER(a.account_type) NOT IN ({noncost}))
        ORDER BY pt.name, l.amount_minor""", list(NONCOST_TYPES)).fetchall()


def compute_audit(conn):
    lines = _cost_lines(conn)

    # --- 1. possible duplicates ---
    duplicates = []
    by_key = defaultdict(list)              # (vendor, amount, memo) -> [rows]
    by_va = defaultdict(list)               # (vendor, amount) -> [rows]
    for r in lines:
        lid, txn, date, vendor, acc, proj, amt, memo = r
        by_key[(vendor, amt, (memo or "").strip().lower())].append(r)
        by_va[(vendor, amt)].append(r)
    seen_pairs = set()
    for (vendor, amt, memo), rows in by_key.items():
        txns = {r[1] for r in rows}
        if len(txns) >= 2:                  # same vendor+amount+memo on separate bills
            duplicates.append(dict(severity="HIGH", vendor=vendor, amount=amt,
                                   count=len(txns), example=rows[0],
                                   reason="same vendor, amount AND description on separate bills"))
            seen_pairs.add((vendor, amt))
    for (vendor, amt), rows in by_va.items():
        if (vendor, amt) in seen_pairs:
            continue
        # same vendor+amount within a short window, even if memo differs
        dates = sorted(_dt.date.fromisoformat(r[2]) for r in rows if r[2])
        txns = {r[1] for r in rows}
        if len(txns) >= 2 and dates and (dates[-1] - dates[0]).days <= DUP_WINDOW_DAYS:
            duplicates.append(dict(severity="MEDIUM", vendor=vendor, amount=amt,
                                   count=len(txns), example=rows[0],
                                   reason=f"same vendor + amount on {len(txns)} bills within {DUP_WINDOW_DAYS} days"))

    # --- 2. unusually large lines per vendor (robust median/MAD) ---
    spikes = []
    per_vendor = defaultdict(list)
    for r in lines:
        per_vendor[r[3]].append(r)
    for vendor, rows in per_vendor.items():
        if len(rows) < SPIKE_MIN_SAMPLE:
            continue
        amts = [r[6] for r in rows]
        med = _median(amts)
        mad = _median([abs(a - med) for a in amts])
        if mad <= 0:
            continue
        for r in rows:
            z = 0.6745 * (r[6] - med) / mad
            if z >= SPIKE_MOD_Z:
                spikes.append(dict(vendor=vendor, example=r, amount=r[6],
                                   median=med, mod_z=round(z, 1),
                                   reason=f"{r[6]/med:.1f}× this vendor's typical line ({fmt_minor(int(med))})"))

    # --- 3. vendor concentration ---
    spend = defaultdict(int)
    for r in lines:
        spend[r[3]] += r[6]
    total = sum(spend.values()) or 1
    concentration = [dict(vendor=v, total=s, pct=100 * s / total)
                     for v, s in spend.items() if s >= CONCENTRATION_MIN]
    concentration.sort(key=lambda x: -x["total"])

    return dict(duplicates=duplicates, spikes=spikes, concentration=concentration,
                total_spend=total, n_lines=len(lines))


def build_expense_audit(db: str, out_xlsx: str) -> dict:
    conn = sqlite3.connect(db)
    conn.execute("PRAGMA foreign_keys = ON")
    meta = dict(conn.execute("SELECT key,value FROM meta").fetchall())
    currency = meta.get("functional_currency", "AUD")
    a = compute_audit(conn)
    conn.close()

    wb = Workbook()
    white = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="2F5496")
    high = PatternFill("solid", fgColor="F8CBAD"); med = PatternFill("solid", fgColor="FFF2CC")
    right = Alignment(horizontal="right")

    def money(ws, r, col, minor):
        c = ws.cell(row=r, column=col, value=float(from_minor(minor)))
        c.number_format = _AUD; c.alignment = right

    # --- Review Alerts (duplicates) ---
    ws = wb.active; ws.title = "Duplicate Alerts"
    ws.append([f"Possible Duplicate Bills — {meta.get('tenant_id')} — REVIEW ALERTS (not confirmed)"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append(["Severity", "Vendor", "Amount", "# Bills", "Example Date", "Example Memo", "Why flagged"])
    for c in ws[2]:
        c.font = white; c.fill = fill
    for d in sorted(a["duplicates"], key=lambda x: (x["severity"] != "HIGH", -x["amount"])):
        ex = d["example"]
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=d["severity"])
        ws.cell(row=r, column=2, value=d["vendor"])
        money(ws, r, 3, d["amount"])
        ws.cell(row=r, column=4, value=d["count"])
        ws.cell(row=r, column=5, value=ex[2])
        ws.cell(row=r, column=6, value=(ex[7] or "")[:60])
        ws.cell(row=r, column=7, value=d["reason"])
        for c in ws[r]:
            c.fill = high if d["severity"] == "HIGH" else med
    for col, w in zip("ABCDEFG", (10, 24, 14, 8, 14, 44, 46)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"

    # --- Unusual cost spikes ---
    ws2 = wb.create_sheet("Unusual Costs")
    ws2.append([f"Unusually Large Costs (robust median/MAD, min {SPIKE_MIN_SAMPLE} samples) — REVIEW ALERTS"])
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.append(["Vendor", "Amount", "Vendor Typical", "× Typical", "Date", "Project", "Account", "Memo"])
    for c in ws2[2]:
        c.font = white; c.fill = fill
    for s in sorted(a["spikes"], key=lambda x: -x["amount"]):
        ex = s["example"]
        r = ws2.max_row + 1
        ws2.cell(row=r, column=1, value=s["vendor"])
        money(ws2, r, 2, s["amount"]); money(ws2, r, 3, int(s["median"]))
        ws2.cell(row=r, column=4, value=round(s["amount"] / s["median"], 1) if s["median"] else "")
        ws2.cell(row=r, column=5, value=ex[2])
        ws2.cell(row=r, column=6, value=ex[5] or "")
        ws2.cell(row=r, column=7, value=ex[4] or "")
        ws2.cell(row=r, column=8, value=(ex[7] or "")[:50])
    for col, w in zip("ABCDEFGH", (24, 14, 16, 10, 12, 10, 18, 46)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A3"

    # --- Vendor concentration ---
    ws3 = wb.create_sheet("Vendor Concentration")
    ws3.append([f"Vendor Concentration — supply-agreement / bulk-buy candidates "
                f"(≥ {fmt_minor(CONCENTRATION_MIN, currency)} spend)"])
    ws3["A1"].font = Font(bold=True, size=13)
    ws3.append(["Vendor", "Total Spend", "% of All Spend"])
    for c in ws3[2]:
        c.font = white; c.fill = fill
    for cn in a["concentration"]:
        r = ws3.max_row + 1
        ws3.cell(row=r, column=1, value=cn["vendor"])
        money(ws3, r, 2, cn["total"])
        ws3.cell(row=r, column=3, value=f"{cn['pct']:.1f}%").alignment = right
    for col, w in zip("ABC", (28, 16, 16)):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A3"

    wb.save(out_xlsx)
    print(f"[audit] {a['n_lines']} cost lines · {len(a['duplicates'])} duplicate alerts · "
          f"{len(a['spikes'])} unusual-cost alerts · {len(a['concentration'])} concentration flags")
    print(f"[audit] wrote {out_xlsx}")
    return a
