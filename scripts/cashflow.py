"""Cash flow forecast (Phase 5) — deterministic, weekly, 90 days, 3 scenarios.

The forecast is arithmetic, not an LLM guess. It projects each project's remaining
contract work and cost-to-complete across its remaining schedule, times the money
in/out by the contract and supplier payment terms, and layers on retention
releases, payroll, overhead and the next GST payment.

Blocking inputs (no trusted business forecast without them): opening cash,
forecast start date, functional currency, payroll, overhead. If any are missing
the forecast runs in PARTIAL mode and says so.

Three scenarios differ ONLY in when client money arrives:
  contractual   payment on the contractual due date
  expected      due date + the client's median days late
  conservative  due date + ~2× median days late (proxy for the slow tail)

Because a QBO transaction-detail export carries no open AR/AP ledger, inflows and
outflows are modelled from remaining contract work + schedule, not from open
invoices. This is disclosed on the Assumptions sheet — the numbers are only as
good as that model, and the report says so.

    build_cashflow_forecast(db, out_xlsx) -> summary
"""

from __future__ import annotations

import datetime as _dt
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from money import from_minor, fmt_minor, to_minor
from jobcost import _approved_budget, NONCOST_TYPES
from retention import compute_retention

HORIZON_WEEKS = 13          # ~90 days
SUPPLIER_TERMS_DAYS = 30    # when we pay our own costs out
_AUD = '#,##0.00'


def _d(iso):
    try:
        return _dt.date.fromisoformat(str(iso)[:10])
    except (ValueError, TypeError):
        return None


def _meta(conn):
    return dict(conn.execute("SELECT key,value FROM meta").fetchall())


def _project_actual(conn, pid):
    noncost = ",".join("?" * len(NONCOST_TYPES))
    return conn.execute(f"""SELECT COALESCE(SUM(l.amount_minor),0)
        FROM transaction_lines l JOIN transaction_headers h ON h.id=l.header_id
        LEFT JOIN accounts a ON a.id=l.account_id
        WHERE l.project_id=? AND l.is_tax=0 AND h.is_void=0
          AND (a.account_type IS NULL OR LOWER(a.account_type) NOT IN ({noncost}))""",
        [pid] + list(NONCOST_TYPES)).fetchone()[0]


def _approved_total(conn, pid):
    return conn.execute("""SELECT COALESCE(SUM(bl.amount_minor),0) FROM budget_lines bl
        JOIN budget_versions bv ON bv.id=bl.budget_version_id
        WHERE bv.project_id=? AND bv.status='approved' AND bv.version_no=0""", (pid,)).fetchone()[0]


def _in_week(start, date):
    """Arrival week for an INFLOW. Overdue money (past date) is still collectible,
    so it lands in week 0; anything past the horizon is out of plan."""
    if not date:
        return None
    w = (date - start).days // 7
    if w < 0:
        return 0
    return w if w < HORIZON_WEEKS else None


def _out_week(start, date):
    """Week for an OUTFLOW. A past-dated payment has already happened, so it is
    dropped from a forward plan; only in-horizon future payments count."""
    if not date:
        return None
    w = (date - start).days // 7
    return w if 0 <= w < HORIZON_WEEKS else None


def _current_approved_total(conn, pid):
    """Sum of the highest APPROVED budget version per cost code (not just v0)."""
    codes = conn.execute("""SELECT DISTINCT cost_code_id FROM budget_lines bl
        JOIN budget_versions bv ON bv.id=bl.budget_version_id WHERE bv.project_id=?""", (pid,)).fetchall()
    return sum(_approved_budget(conn, pid, ccid, None) for (ccid,) in codes)


def compute_cashflow(conn):
    meta = _meta(conn)
    start = _d(meta.get("cash_forecast_start_date"))
    opening = meta.get("cash_opening_bank_balance")
    payroll_fn = meta.get("cash_payroll_fortnightly")
    overhead_m = meta.get("cash_overhead_monthly")
    missing = [k for k, v in [("opening_bank_balance", opening),
               ("forecast_start_date", start), ("payroll_fortnightly", payroll_fn),
               ("overhead_monthly", overhead_m)] if not v]
    if start is None:
        start = _dt.date.today()

    opening_minor = to_minor(str(opening)) if opening else 0
    payroll_minor = to_minor(str(payroll_fn)) if payroll_fn else 0
    overhead_wk = to_minor(str(overhead_m)) * 12 // 52 if overhead_m else 0
    gst_date = _d(meta.get("cash_gst_next_payment_date"))

    ret = {r["code"]: r for r in compute_retention(conn, as_of=None)}

    scenarios = ("contractual", "expected", "conservative")
    # weekly line items per scenario
    inflow = {s: [0] * HORIZON_WEEKS for s in scenarios}      # progress claims / receivables
    releases = {s: [0] * HORIZON_WEEKS for s in scenarios}    # retention released to us
    cost_out = [0] * HORIZON_WEEKS                            # supplier/subcontractor costs
    sub_ret_out = [0] * HORIZON_WEEKS                         # retention we release to subs

    # INVOICE-BASED mode when A/R or A/P aging has been ingested — real open
    # invoices with due dates, not a schedule estimate.
    has_ar = conn.execute("SELECT COUNT(*) FROM ar_items").fetchone()[0]
    has_ap = conn.execute("SELECT COUNT(*) FROM ap_items").fetchone()[0]
    mode = "invoice-based" if (has_ar or has_ap) else "schedule-based"

    if mode == "invoice-based":
        # receivables in: due date + scenario delay (a debtor's median lateness)
        med_by_client = dict(conn.execute("""SELECT ct.party_id, ct.median_days_late
            FROM contracts ct""").fetchall())
        for party_id, due, open_minor in conn.execute(
                "SELECT party_id, due_date, open_minor FROM ar_items WHERE open_minor != 0"):
            med = med_by_client.get(party_id, 0) or 0
            for s, delay in (("contractual", 0), ("expected", med), ("conservative", med * 2)):
                iw = _in_week(start, _d(due))
                if iw is not None:
                    # apply the scenario delay by shifting the due date
                    iw2 = _in_week(start, _d(due) + _dt.timedelta(days=delay)) if _d(due) else None
                    if iw2 is not None:
                        inflow[s][iw2] += open_minor
        # payables out: on the due date (we pay when due)
        for due, open_minor in conn.execute(
                "SELECT due_date, open_minor FROM ap_items WHERE open_minor != 0"):
            ow = _out_week(start, _d(due))
            if ow is not None:
                cost_out[ow] += open_minor

    for pid, code, status in conn.execute("SELECT id, code, status FROM projects"):
        c = conn.execute("""SELECT contract_value_minor, payment_terms_days, retention_pct_bp,
            median_days_late FROM contracts WHERE project_id=? LIMIT 1""", (pid,)).fetchone()
        if not c:
            continue
        cval, terms, rbp, med_late = c
        terms = terms or 30
        med_late = med_late or 0
        finish = conn.execute("SELECT MAX(planned_finish) FROM activities WHERE project_id=?", (pid,)).fetchone()[0]
        fin = _d(finish)
        rrow = ret.get(code, {})

        # Schedule-based projection of billing + cost — only when we DON'T have
        # real open invoices. With A/R and A/P loaded, actual invoices drive the
        # money in/out and this estimate is skipped.
        if mode == "schedule-based":
            revenue_certified = rrow.get("revenue_certified", 0)
            remaining_bill = max((cval or 0) - revenue_certified, 0)
            remaining_cost = max(_current_approved_total(conn, pid) - _project_actual(conn, pid), 0)
            rem_weeks = 0
            if fin and fin > start:
                rem_weeks = max((fin - start).days // 7, 1)
            if rem_weeks and (remaining_bill or remaining_cost):
                weekly_bill = remaining_bill // rem_weeks
                weekly_cost = remaining_cost // rem_weeks
                for wk in range(min(rem_weeks, HORIZON_WEEKS)):
                    bill_date = start + _dt.timedelta(weeks=wk)
                    ow = _out_week(start, bill_date + _dt.timedelta(days=SUPPLIER_TERMS_DAYS))
                    if ow is not None:
                        cost_out[ow] += weekly_cost
                    for s, delay in (("contractual", 0), ("expected", med_late),
                                     ("conservative", med_late * 2)):
                        iw = _in_week(start, bill_date + _dt.timedelta(days=terms + delay))
                        if iw is not None:
                            inflow[s][iw] += weekly_bill

        # retention releases (to us) at release date; sub retention (out) at project finish
        # — these apply in BOTH modes (retention isn't an ordinary open invoice)
        rel = _d(rrow.get("release"))
        client_ret = rrow.get("client_retention", 0)
        if rel and client_ret:
            for s, delay in (("contractual", 0), ("expected", med_late), ("conservative", med_late * 2)):
                rw = _in_week(start, rel + _dt.timedelta(days=delay))
                if rw is not None:
                    releases[s][rw] += client_ret
        sub_ret = rrow.get("sub_retention", 0)
        if fin and sub_ret:
            sw = _out_week(start, fin)
            if sw is not None:
                sub_ret_out[sw] += sub_ret

    # fixed outflows
    payroll_wk = [payroll_minor if (w % 2 == 0) else 0 for w in range(HORIZON_WEEKS)]  # fortnightly
    overhead = [overhead_wk] * HORIZON_WEEKS
    gst = [0] * HORIZON_WEEKS
    gw = _out_week(start, gst_date) if gst_date else None
    gst_amt = to_minor(str(meta.get("cash_gst_next_payment_amount"))) if meta.get("cash_gst_next_payment_amount") else 0
    if gw is not None:
        gst[gw] += gst_amt

    # assemble running balances
    weeks = [(start + _dt.timedelta(weeks=w)).isoformat() for w in range(HORIZON_WEEKS)]
    result = dict(start=start.isoformat(), weeks=weeks, opening=opening_minor,
                  missing=missing, mode=mode, scenarios={})
    for s in scenarios:
        bal = opening_minor
        rows = []
        for w in range(HORIZON_WEEKS):
            inflow_w = inflow[s][w] + releases[s][w]
            outflow_w = cost_out[w] + sub_ret_out[w] + payroll_wk[w] + overhead[w] + gst[w]
            net = inflow_w - outflow_w
            bal += net
            rows.append(dict(week=weeks[w], claims=inflow[s][w], releases=releases[s][w],
                             cost=cost_out[w], sub_ret=sub_ret_out[w], payroll=payroll_wk[w],
                             overhead=overhead[w], gst=gst[w], net=net, balance=bal))
        low = min(rows, key=lambda r: r["balance"])
        shortfall = next((r["week"] for r in rows if r["balance"] < 0), None)
        result["scenarios"][s] = dict(rows=rows, min_balance=low["balance"],
                                      min_week=low["week"], first_shortfall=shortfall)
    return result


def build_cashflow_forecast(db: str, out_xlsx: str) -> dict:
    conn = sqlite3.connect(db)
    conn.execute("PRAGMA foreign_keys = ON")
    meta = _meta(conn)
    currency = meta.get("functional_currency", "AUD")
    cf = compute_cashflow(conn)
    conn.close()

    partial = bool(cf["missing"])
    wb = Workbook()

    # --- Summary ---
    ws = wb.active; ws.title = "Cash Flow Summary"
    white = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="2F5496")
    bad = PatternFill("solid", fgColor="F8CBAD"); good = PatternFill("solid", fgColor="E2EFDA")
    right = Alignment(horizontal="right")
    invoice_based = cf["mode"] == "invoice-based"
    label = "13-Week Cash Flow Forecast" if invoice_based else "13-Week Schedule-Based Cash Plan"
    ws.append([("PARTIAL — missing: " + ", ".join(cf["missing"])) if partial
               else f"{label} — {meta.get('tenant_id')}"])
    ws["A1"].font = Font(bold=True, size=13, color="C00000" if partial else "000000")
    ws.append([f"13 weeks from {cf['start']} · opening balance {fmt_minor(cf['opening'], currency)} · "
               f"{currency} · " + ("driven by real open A/R + A/P invoices" if invoice_based
               else "schedule-based estimate (no open AR/AP loaded)") + " — see Assumptions"])
    ws.append([])
    ws.append(["Scenario", "Lowest Balance", "In Week", "First Cash Shortfall"])
    for c in ws[ws.max_row]:
        c.font = white; c.fill = fill
    for s in ("contractual", "expected", "conservative"):
        sc = cf["scenarios"][s]
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=s)
        mc = ws.cell(row=r, column=2, value=float(from_minor(sc["min_balance"])))
        mc.number_format = _AUD; mc.alignment = right
        ws.cell(row=r, column=3, value=sc["min_week"])
        ws.cell(row=r, column=4, value=sc["first_shortfall"] or "none")
        if sc["min_balance"] < 0:
            for c in ws[r]:
                c.fill = bad
        elif s == "expected":
            for c in ws[r]:
                c.fill = good
    for i, w in enumerate([16, 18, 14, 20], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # --- Weekly detail (expected scenario, headline) ---
    ws2 = wb.create_sheet("Weekly (Expected)")
    line_hdrs = ["Week starting", "Progress Claims In", "Retention Released In",
                 "Subcontractor/Supplier Cost", "Sub Retention Paid", "Payroll",
                 "Overhead", "GST", "Net Movement", "Closing Balance"]
    ws2.append(line_hdrs)
    for c in ws2[1]:
        c.font = white; c.fill = fill
    for row in cf["scenarios"]["expected"]["rows"]:
        r = ws2.max_row + 1
        ws2.cell(row=r, column=1, value=row["week"])
        vals = [row["claims"], row["releases"], -row["cost"], -row["sub_ret"],
                -row["payroll"], -row["overhead"], -row["gst"], row["net"], row["balance"]]
        for i, v in enumerate(vals, 2):
            cell = ws2.cell(row=r, column=i, value=float(from_minor(v)))
            cell.number_format = _AUD; cell.alignment = right
        if row["balance"] < 0:
            for c in ws2[r]:
                c.fill = bad
    for i, w in enumerate([16, 18, 20, 26, 18, 14, 14, 12, 16, 18], 1):
        ws2.column_dimensions[chr(64 + i)].width = w
    ws2.freeze_panes = "A2"

    # --- Assumptions (disclosure) ---
    ws3 = wb.create_sheet("Assumptions")
    ws3.append(["Assumptions & coverage"])
    ws3["A1"].font = Font(bold=True, size=13)
    notes = [
        ("Horizon", f"{HORIZON_WEEKS} weeks from {cf['start']}"),
        ("Mode", ("invoice-based — real open A/R and A/P drive money in/out"
                  if cf["mode"] == "invoice-based" else
                  "schedule-based estimate — no open AR/AP loaded")),
        ("Opening balance", fmt_minor(cf["opening"], currency)),
        ("Scenarios", "contractual = due date; expected = +median days late; conservative = +2× median"),
        ("Inflows", ("open receivables (A/R) by due date + client's median lateness"
                     if cf["mode"] == "invoice-based" else
                     "remaining contract-to-bill spread over schedule, timed by payment terms")),
        ("Outflows", ("open payables (A/P) by due date"
                      if cf["mode"] == "invoice-based" else
                      f"remaining cost-to-complete, paid {SUPPLIER_TERMS_DAYS}d after incurred")
                     + "; + payroll, overhead, GST"),
        ("Not-yet-invoiced work", ("excluded — only issued invoices are forecast (conservative)"
                                   if cf["mode"] == "invoice-based" else "projected from schedule")),
        ("Retention", "client retention released at planned trigger date; overdue retention shown in week 1; sub retention paid at project finish"),
        ("Cost-to-complete", "current approved budget − actual; excludes forecast overruns and open commitments beyond budget"),
        ("Billing", "remaining revenue spread evenly per week (claim frequency not applied); costs paid 30d after incurred (generic, not per-vendor terms)"),
        ("NOT modelled", "open AR/AP invoice ledger; actual certified claims; retention already released — none are in a QBO transaction-detail export"),
        ("Payroll", f"{fmt_minor(to_minor(str(meta.get('cash_payroll_fortnightly') or 0)), currency)} fortnightly"),
        ("Overhead", f"{fmt_minor(to_minor(str(meta.get('cash_overhead_monthly') or 0)), currency)} monthly, spread weekly"),
        ("GST", (f"payment {meta.get('cash_gst_next_payment_date')}, "
                 + (fmt_minor(to_minor(str(meta.get('cash_gst_next_payment_amount'))), currency)
                    if meta.get('cash_gst_next_payment_amount') else "amount NOT provided — excluded from the plan"))),
        ("Blocking inputs", ", ".join(cf["missing"]) or "all present (schedule-based plan; not a full treasury forecast)"),
    ]
    for k, v in notes:
        ws3.append([k, v])
    ws3.column_dimensions["A"].width = 26; ws3.column_dimensions["B"].width = 70

    wb.save(out_xlsx)

    exp = cf["scenarios"]["expected"]
    con = cf["scenarios"]["conservative"]
    print(f"[cashflow] {'PARTIAL' if partial else cf['mode']} · 13wk from {cf['start']} · opening {fmt_minor(cf['opening'], currency)}")
    print(f"[cashflow] expected: low {fmt_minor(exp['min_balance'], currency)} in wk {exp['min_week']} · "
          f"shortfall {exp['first_shortfall'] or 'none'}")
    print(f"[cashflow] conservative: low {fmt_minor(con['min_balance'], currency)} · "
          f"shortfall {con['first_shortfall'] or 'none'}")
    print(f"[cashflow] wrote {out_xlsx}")
    return cf
