"""Cost-code classification (Phase 3, hardened).

Method stack, per uncoded cost line (brief §6):
    1. rule        approved mapping rule, PROJECT-scoped, conflict-safe   -> auto-apply
    2. learned     approved rule promoted from prior reviews             -> auto-apply
    3. commitment  unique (project, vendor) subcontract commitment       -> PROPOSAL
    4. llm/heuristic injected proposer / token-overlap fallback          -> PROPOSAL
    5. unclassified explicit abstain                                     -> stays uncoded

Only APPROVED rules auto-apply. Everything else — including commitment matches —
is a proposal that waits for human sign-off. Nothing is force-fit; proposals are
validated against the line's PROJECT allowed cost-code list (not the whole tenant
chart), so Claude can never introduce a code that isn't valid for that project.

Safety around persistence (independent review):
  * Review workbooks are bound to this database (db_uid) + tenant + a batch token.
    A workbook from another db cannot modify these records.
  * Applying a workbook is idempotent: UNIQUE(review_batch, classification) plus an
    optimistic-concurrency check on a per-line content hash. Re-applying, or
    applying a stale workbook after the line moved, makes no changes.
  * Rule promotion is conservative: a single acceptance never creates an
    auto-applying rule. Rules are 'candidate' until they have >=2 consistent,
    conflict-free confirmations OR an explicit "approve rule" decision.

    classify_db(db, proposer=None) -> stats
    write_review_workbook(db, out_xlsx) -> batch_token
    apply_review(db, review_xlsx) -> result buckets
"""

from __future__ import annotations

import datetime as _dt
import hashlib
import re
import sqlite3
import uuid
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation

from money import from_minor

REVIEW_SCHEMA_VERSION = "review-0.2.0"
HEURISTIC_FLOOR_BP = 3000
PROMOTE_MIN_SUPPORT = 2       # consistent confirmations before a candidate rule forms

NONCOST_TYPES = ("income", "revenue", "asset", "liability", "equity",
                 "bank", "accounts receivable", "accounts payable")

_STOP = set("""the a an and or of for to in on at by with from into over per
 project stage services service works work package packages supply install
 installation progress claim invoice bill charge jan feb mar apr may jun jul aug
 sep oct nov dec 2024 2025 2026 extension refurbishment fitout warehouse centre
 apartment tower medical logistics ridge summit""".split())


def _now() -> str:
    return _dt.datetime.now().isoformat(timespec="seconds")


def _tok(s: str) -> set:
    return {t for t in re.split(r"[^a-z0-9]+", (s or "").lower())
            if t and t not in _STOP and len(t) > 2}


def _line_hash(txn, account, project, amount_minor, memo) -> str:
    return hashlib.sha256(
        f"{txn}|{account}|{project}|{amount_minor}|{memo}".encode("utf-8")).hexdigest()[:16]


# --------------------------------------------------------------------------- #
# reference / matchers
# --------------------------------------------------------------------------- #

def _allowed_codes(conn, tenant):
    rows = conn.execute("SELECT id, code, name FROM cost_codes WHERE tenant_id=?",
                        (tenant,)).fetchall()
    return [dict(id=i, code=c, name=n, tokens=_tok(n)) for i, c, n in rows]


def _project_allowed(conn):
    """{project_id: set(cost_code_id)} — the codes actually valid on each project
    (its activities). Classification is bound to this, not the whole tenant chart."""
    m = defaultdict(set)
    for pid, ccid in conn.execute(
            "SELECT project_id, cost_code_id FROM activities WHERE cost_code_id IS NOT NULL"):
        m[pid].add(ccid)
    return m


def _match_rule(conn, tenant, party_id, account_id, project_id):
    """Most-specific APPROVED rule. Deterministic; if the most specific tier holds
    two rules with different codes, abstain (return None) instead of guessing."""
    rows = conn.execute("""
        SELECT cost_code_id, party_id, account_id, project_id, created_from_review_id
        FROM mapping_rules WHERE tenant_id=? AND status='approved'
        ORDER BY id""", (tenant,)).fetchall()
    matches = []
    for code_id, rp, ra, rpr, from_rev in rows:
        if rp is not None and rp != party_id:   continue
        if ra is not None and ra != account_id: continue
        if rpr is not None and rpr != project_id: continue
        spec = (rp is not None) + (ra is not None) + (rpr is not None)
        matches.append((spec, code_id, from_rev is not None))
    if not matches:
        return None, False
    top = max(m[0] for m in matches)
    top_codes = {m[1] for m in matches if m[0] == top}
    if len(top_codes) != 1:
        return None, False           # conflicting rules at the same specificity -> abstain
    winner = next(m for m in matches if m[0] == top)
    return winner[1], winner[2]


def _match_commitment(conn, project_id, party_id):
    if party_id is None:
        return None
    rows = conn.execute("""SELECT DISTINCT cost_code_id FROM commitments
        WHERE project_id=? AND party_id=? AND cost_code_id IS NOT NULL""",
        (project_id, party_id)).fetchall()
    return rows[0][0] if len(rows) == 1 else None


def _heuristic(line_tokens, allowed, allowed_ids):
    best_id, best = None, 0
    for cc in allowed:
        if cc["id"] not in allowed_ids or not cc["tokens"]:
            continue
        inter = line_tokens & cc["tokens"]
        if not inter:
            continue
        score = int(10000 * len(inter) / len(cc["tokens"]))
        if score > best:
            best_id, best = cc["id"], score
    return (best_id, best) if best >= HEURISTIC_FLOOR_BP else (None, 0)


def _candidates(conn):
    noncost = ",".join("?" * len(NONCOST_TYPES))
    return conn.execute(f"""
        SELECT l.id, h.native_transaction_id, l.project_id, l.account_id, h.party_id,
               a.name, pt.name, p.code, l.amount_minor, l.memo
        FROM transaction_lines l
        JOIN transaction_headers h ON h.id = l.header_id
        LEFT JOIN accounts a ON a.id = l.account_id
        LEFT JOIN parties pt ON pt.id = h.party_id
        LEFT JOIN projects p ON p.id = l.project_id
        WHERE l.cost_code_id IS NULL AND l.project_id IS NOT NULL
          AND l.is_tax=0 AND h.is_void=0
          AND (a.account_type IS NULL OR LOWER(a.account_type) NOT IN ({noncost}))
        ORDER BY l.id""", list(NONCOST_TYPES)).fetchall()


# --------------------------------------------------------------------------- #
# classify
# --------------------------------------------------------------------------- #

def classify_db(db: str, proposer=None) -> dict:
    conn = sqlite3.connect(db)
    conn.execute("PRAGMA foreign_keys = ON")
    tenant = dict(conn.execute("SELECT key,value FROM meta").fetchall())["tenant_id"]
    allowed = _allowed_codes(conn, tenant)
    code_by_str = {c["code"]: c["id"] for c in allowed}
    proj_allowed = _project_allowed(conn)

    # idempotent re-run: drop stale classifications for still-uncoded lines
    conn.execute("""DELETE FROM classifications WHERE transaction_line_id IN
        (SELECT id FROM transaction_lines WHERE cost_code_id IS NULL)""")

    stats = defaultdict(int)
    coded_val = total_val = 0
    for (lid, txn, pid, aid, party_id, acc, vendor, pcode, amt, memo) in _candidates(conn):
        total_val += amt
        allowed_ids = proj_allowed.get(pid) or {c["id"] for c in allowed}
        method = code_id = score = reason = None

        rule_code, learned = _match_rule(conn, tenant, party_id, aid, pid)
        if rule_code is not None and rule_code in allowed_ids:
            method, code_id = ("learned" if learned else "rule"), rule_code
            reason = "approved mapping rule"
        if code_id is None:
            cm = _match_commitment(conn, pid, party_id)
            if cm is not None and cm in allowed_ids:
                method, code_id, score, reason = "commitment", cm, 9000, \
                    "unique subcontract commitment for this vendor+project (proposal)"
        if code_id is None:
            if proposer is not None:
                prop = proposer(dict(memo=memo, vendor=vendor, account=acc, project=pcode,
                                     amount=float(from_minor(amt))),
                                [c for c in allowed if c["id"] in allowed_ids]) or {}
                cand = prop.get("cost_code")
                if cand in code_by_str and code_by_str[cand] in allowed_ids:
                    method, code_id = "llm", code_by_str[cand]
                    score, reason = prop.get("score"), prop.get("reason", "llm proposal")
            else:
                hid, hs = _heuristic(_tok(memo) | _tok(vendor), allowed, allowed_ids)
                if hid is not None:
                    method, code_id, score, reason = "heuristic", hid, hs, "token overlap with cost-code name"

        if code_id is None:
            method, reason = "unclassified", "no rule, and no confident project-valid proposal"

        conn.execute("INSERT INTO classifications(transaction_line_id, proposed_cost_code_id, "
                     "method, score, reason, created_at) VALUES (?,?,?,?,?,?)",
                     (lid, code_id, method, score, reason, _now()))
        if method in ("rule", "learned"):          # ONLY approved rules auto-apply
            conn.execute("UPDATE transaction_lines SET cost_code_id=? WHERE id=?", (code_id, lid))
            coded_val += amt
        stats[method] += 1

    conn.commit()
    conn.close()
    total = sum(stats.values())
    auto = stats["rule"] + stats["learned"]
    line_cov = (100 * auto / total) if total else 100.0
    val_cov = (100 * coded_val / total_val) if total_val else 100.0
    print(f"[classify] {total} uncoded lines · auto-applied {auto} (approved rules only: "
          f"rule={stats['rule']} learned={stats['learned']}) · "
          f"proposals: commitment={stats['commitment']} heuristic={stats['heuristic']} llm={stats['llm']} · "
          f"unclassified={stats['unclassified']}")
    print(f"[classify] auto-coded coverage — lines {line_cov:.2f}% · value {val_cov:.2f}%")
    return dict(stats)


# --------------------------------------------------------------------------- #
# review workbook (batch-bound; rule-group + line sheets)
# --------------------------------------------------------------------------- #

_LINE_HDRS = ["classification_id", "transaction_line_id", "exported_code", "method",
              "score %", "line_hash", "date", "project", "account", "vendor",
              "amount", "memo", "decision (accept/replace)", "replacement_code", "comment"]
_GROUP_HDRS = ["group_id", "vendor", "account", "project", "proposed_code", "method",
               "lines", "total_value", "min_score %", "example_memos",
               "decision (approve rule/review lines)"]


def _san(v):
    return "'" + v if isinstance(v, str) and v[:1] in ("=", "+", "-", "@") else v


def write_review_workbook(db: str, out_xlsx: str) -> str:
    conn = sqlite3.connect(db)
    meta = dict(conn.execute("SELECT key,value FROM meta").fetchall())
    tenant, db_uid = meta["tenant_id"], meta["db_uid"]

    rows = conn.execute("""
        SELECT c.id, c.transaction_line_id, cc.code, c.method, c.score,
               h.native_transaction_id, h.txn_date, p.code, a.name, pt.name,
               l.amount_minor, l.memo, h.party_id, l.account_id, l.project_id, cc.id
        FROM classifications c
        JOIN transaction_lines l ON l.id = c.transaction_line_id
        JOIN transaction_headers h ON h.id = l.header_id
        LEFT JOIN cost_codes cc ON cc.id = c.proposed_cost_code_id
        LEFT JOIN projects p ON p.id = l.project_id
        LEFT JOIN accounts a ON a.id = l.account_id
        LEFT JOIN parties pt ON pt.id = h.party_id
        WHERE c.method IN ('llm','heuristic','commitment','unclassified')
          AND l.cost_code_id IS NULL
        ORDER BY c.id""").fetchall()

    token = uuid.uuid4().hex
    batch_id = conn.execute(
        "INSERT INTO review_batches(tenant_id, db_uid, batch_token, created_at, row_count, status) "
        "VALUES (?,?,?,?,?, 'open')", (tenant, db_uid, token, _now(), len(rows))).lastrowid
    conn.commit()

    # allowed codes per project (reference + validation)
    proj_codes = defaultdict(list)
    for pcode, code, name in conn.execute("""
            SELECT p.code, cc.code, cc.name FROM activities ac
            JOIN projects p ON p.id=ac.project_id JOIN cost_codes cc ON cc.id=ac.cost_code_id
            ORDER BY p.code, cc.code"""):
        proj_codes[pcode].append((code, name))
    all_codes = sorted({c for lst in proj_codes.values() for c, _ in lst})
    conn.close()

    wb = Workbook()

    # --- Batch sheet (binding) ---
    bs = wb.active; bs.title = "Batch"
    bs.append(["Construction CFO — classification review batch"])
    bs["A1"].font = Font(bold=True, size=13)
    for k, v in [("schema", REVIEW_SCHEMA_VERSION), ("tenant_id", tenant),
                 ("db_uid", db_uid), ("batch_token", token), ("row_count", len(rows)),
                 ("exported_at", _now())]:
        bs.append([k, _san(str(v))])
    bs.append([])
    bs.append(["How to review:"])
    bs.append(["1) Prefer the 'Rule Groups' sheet — approve a whole vendor→code mapping in one row."])
    bs.append(["2) Use 'Line Review' for exceptions. decision = accept or replace (+ replacement_code)."])
    bs.append(["3) Do NOT edit id / hash columns. Save and run: cfo review-apply."])
    bs.column_dimensions["A"].width = 16; bs.column_dimensions["B"].width = 44

    lock = Protection(locked=True); unlock = Protection(locked=False)
    hfont = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor="2F5496")

    # --- Rule Groups sheet ---
    gs = wb.create_sheet("Rule Groups")
    groups = defaultdict(lambda: dict(lines=0, value=0, min_score=None, memos=[], method=set()))
    for (cid, lid, code, method, score, txn, date, pcode, acc, vendor, amt, memo,
         party_id, acc_id, proj_id, ccid) in rows:
        if code is None:
            continue                         # abstentions handled per-line only
        key = (vendor or "", acc or "", pcode or "", code)
        g = groups[key]
        g["lines"] += 1; g["value"] += amt; g["method"].add(method)
        if score is not None:
            g["min_score"] = score if g["min_score"] is None else min(g["min_score"], score)
        if len(g["memos"]) < 3 and memo:
            g["memos"].append(memo[:40])
    gs.append(_GROUP_HDRS)
    for c in gs[1]:
        c.font = hfont; c.fill = hfill
    for gid, ((vendor, acc, pcode, code), g) in enumerate(sorted(groups.items()), 1):
        ms = "" if g["min_score"] is None else f'{g["min_score"]/100:.0f}%'
        gs.append([gid, _san(vendor), _san(acc), _san(pcode), _san(code),
                   "/".join(sorted(g["method"])), g["lines"], float(from_minor(g["value"])),
                   ms, _san(" | ".join(g["memos"])), ""])
        gs.cell(row=gs.max_row, column=8).number_format = '#,##0.00'
        gs.cell(row=gs.max_row, column=11).protection = unlock
    dv_g = DataValidation(type="list", formula1='"approve rule,review lines,reject"', allow_blank=True)
    gs.add_data_validation(dv_g); dv_g.add(f"K2:K{max(gs.max_row,2)}")
    gs.protection.sheet = True; gs.protection.enable()
    for col, w in zip("ABCDEFGHIJK", (8, 22, 16, 10, 12, 14, 7, 14, 10, 44, 26)):
        gs.column_dimensions[col].width = w
    gs.freeze_panes = "A2"

    # --- Line Review sheet ---
    ls = wb.create_sheet("Line Review")
    ls.append(_LINE_HDRS)
    for c in ls[1]:
        c.font = hfont; c.fill = hfill
    for (cid, lid, code, method, score, txn, date, pcode, acc, vendor, amt, memo,
         party_id, acc_id, proj_id, ccid) in rows:
        lh = _line_hash(txn, acc, pcode, amt, memo)
        r = [cid, lid, _san(code or "UNCLASSIFIED"), method,
             "" if score is None else f"{score/100:.0f}%", lh, _san(date or ""),
             _san(pcode or ""), _san(acc or ""), _san(vendor or ""),
             float(from_minor(amt)), _san((memo or "")[:80]), "", "", ""]
        ls.append(r)
        row_i = ls.max_row
        ls.cell(row=row_i, column=11).number_format = '#,##0.00'
        for col in (13, 14, 15):                 # decision / replacement / comment editable
            ls.cell(row=row_i, column=col).protection = unlock
    dv_d = DataValidation(type="list", formula1='"accept,replace"', allow_blank=True)
    ls.add_data_validation(dv_d); dv_d.add(f"M2:M{max(ls.max_row,2)}")
    if all_codes:
        dv_c = DataValidation(type="list", formula1='"' + ",".join(all_codes[:200]) + '"', allow_blank=True)
        ls.add_data_validation(dv_c); dv_c.add(f"N2:N{max(ls.max_row,2)}")
    ls.protection.sheet = True; ls.protection.enable()
    for col, w in zip("ABCDEFGHIJKLMNO", (15, 16, 14, 11, 8, 16, 11, 9, 16, 20, 12, 44, 22, 16, 20)):
        ls.column_dimensions[col].width = w
    ls.freeze_panes = "A2"

    # --- Allowed Codes reference ---
    rs = wb.create_sheet("Allowed Codes")
    rs.append(["project", "cost_code", "name"])
    for c in rs[1]:
        c.font = hfont; c.fill = hfill
    for pcode in sorted(proj_codes):
        for code, name in proj_codes[pcode]:
            rs.append([_san(pcode), _san(code), _san(name)])
    for col, w in zip("ABC", (12, 14, 40)):
        rs.column_dimensions[col].width = w

    wb.save(out_xlsx)
    print(f"[review-export] batch {token[:8]}… · {len(rows)} lines · {len(groups)} rule-groups -> {out_xlsx}")
    return token


# --------------------------------------------------------------------------- #
# apply
# --------------------------------------------------------------------------- #

def _read_batch(ws_batch):
    kv = {}
    for row in ws_batch.iter_rows(values_only=True):
        if row and row[0] and len(row) > 1 and row[1] is not None:
            kv[str(row[0]).strip()] = str(row[1]).strip()
    return kv


def apply_review(db: str, review_xlsx: str) -> dict:
    wb = load_workbook(review_xlsx, data_only=True)
    if "Batch" not in wb.sheetnames:
        raise ValueError("not a valid review workbook (no Batch sheet)")
    b = _read_batch(wb["Batch"])
    if b.get("schema") != REVIEW_SCHEMA_VERSION:
        raise ValueError(f"incompatible review schema {b.get('schema')} != {REVIEW_SCHEMA_VERSION}")

    conn = sqlite3.connect(db)
    conn.execute("PRAGMA foreign_keys = ON")
    meta = dict(conn.execute("SELECT key,value FROM meta").fetchall())
    tenant, db_uid = meta["tenant_id"], meta["db_uid"]
    # batch binding: the workbook must belong to THIS database + tenant + a known batch
    if b.get("db_uid") != db_uid or b.get("tenant_id") != tenant:
        conn.close()
        raise ValueError("review workbook is bound to a different database/tenant")
    batch = conn.execute("SELECT id, status FROM review_batches WHERE batch_token=?",
                         (b.get("batch_token"),)).fetchone()
    if not batch:
        conn.close()
        raise ValueError("unknown batch token — workbook was not exported from this database")
    batch_id = batch[0]

    code_by_str = {c: i for c, i in conn.execute(
        "SELECT code, id FROM cost_codes WHERE tenant_id=?", (tenant,)).fetchall()}
    proj_allowed = _project_allowed(conn)

    res = dict(applied=0, already=0, stale=0, rejected=0, rules_approved=0,
               candidates=0, errors=[])
    support = defaultdict(lambda: defaultdict(int))   # scope -> {code_id: count}

    def line_state(line_id):
        r = conn.execute("""SELECT l.cost_code_id, h.native_transaction_id, a.name, p.code,
            l.amount_minor, l.memo, h.party_id, l.account_id, l.project_id
            FROM transaction_lines l JOIN transaction_headers h ON h.id=l.header_id
            LEFT JOIN accounts a ON a.id=l.account_id LEFT JOIN projects p ON p.id=l.project_id
            WHERE l.id=?""", (line_id,)).fetchone()
        return r

    def do_apply(cid, line_id, code_id, decision, comment, exported_hash):
        st = line_state(line_id)
        if st is None:
            res["rejected"] += 1; res["errors"].append(f"cls {cid}: unknown line"); return
        cur_code, txn, acc, pcode, amt, memo, party_id, acc_id, proj_id = st
        # project-scoped validity
        if code_id not in (proj_allowed.get(proj_id) or set()):
            res["rejected"] += 1
            res["errors"].append(f"cls {cid}: code not valid for project"); return
        # idempotency: a review for (batch, classification) already exists?
        if conn.execute("SELECT 1 FROM classification_reviews WHERE review_batch_id=? AND classification_id=?",
                        (batch_id, cid)).fetchone():
            res["already"] += 1; return
        # optimistic concurrency: line must be unchanged and still uncoded
        if cur_code is not None:
            res["stale"] += 1; return
        if exported_hash and exported_hash != _line_hash(txn, acc, pcode, amt, memo):
            res["stale"] += 1
            res["errors"].append(f"cls {cid}: line changed since export (stale)"); return
        conn.execute("UPDATE transaction_lines SET cost_code_id=? WHERE id=?", (code_id, line_id))
        conn.execute("""INSERT INTO classification_reviews(classification_id, transaction_line_id,
            review_batch_id, original_code_id, reviewer_decision, replacement_code_id, line_hash,
            reviewer_comment, reviewed_at, workbook_version)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (cid, line_id, batch_id,
             conn.execute("SELECT proposed_cost_code_id FROM classifications WHERE id=?", (cid,)).fetchone()[0],
             decision, code_id, exported_hash, comment or "", _now(), REVIEW_SCHEMA_VERSION))
        res["applied"] += 1
        support[(party_id, acc_id, proj_id)][code_id] += 1

    # ---- Rule Groups (explicit approvals) ----
    approved_scopes = set()
    if "Rule Groups" in wb.sheetnames:
        gs = list(wb["Rule Groups"].iter_rows(values_only=True))
        ghdr = [str(c).strip() if c else "" for c in gs[0]]
        gi = {h: k for k, h in enumerate(ghdr)}
        for row in gs[1:]:
            if not row or row[gi["group_id"]] in (None, ""):
                continue
            decision = str(row[gi["decision (approve rule/review lines)"]] or "").strip().lower()
            if decision != "approve rule":
                continue
            vendor = str(row[gi["vendor"]] or "").lstrip("'")
            acc = str(row[gi["account"]] or "").lstrip("'")
            pcode = str(row[gi["project"]] or "").lstrip("'")
            code = str(row[gi["proposed_code"]] or "").lstrip("'")
            if code not in code_by_str:
                res["rejected"] += 1; res["errors"].append(f"group {row[gi['group_id']]}: bad code"); continue
            code_id = code_by_str[code]
            ids = conn.execute("""SELECT p.id, a.id, pt.id FROM projects p, accounts a, parties pt
                WHERE p.code=? AND a.name=? AND pt.name=? AND p.tenant_id=?""",
                (pcode, acc, vendor, tenant)).fetchone()
            if not ids:
                continue
            proj_id, acc_id, party_id = ids
            if code_id not in (proj_allowed.get(proj_id) or set()):
                res["rejected"] += 1; res["errors"].append(f"group: code not valid for {pcode}"); continue
            # explicit approval -> approved rule (idempotent)
            ex = conn.execute("""SELECT id FROM mapping_rules WHERE tenant_id=? AND status='approved'
                AND party_id=? AND account_id=? AND project_id=? AND cost_code_id=?""",
                (tenant, party_id, acc_id, proj_id, code_id)).fetchone()
            if not ex:
                conn.execute("""INSERT INTO mapping_rules(tenant_id, party_id, account_id, project_id,
                    cost_code_id, status, support_count, created_at)
                    VALUES (?,?,?,?,?, 'approved', 0, ?)""",
                    (tenant, party_id, acc_id, proj_id, code_id, _now()))
                res["rules_approved"] += 1
            approved_scopes.add((party_id, acc_id, proj_id))
            # apply the approved rule to every matching uncoded classification in this batch
            for cid, lid, chash in conn.execute("""
                    SELECT c.id, c.transaction_line_id, NULL FROM classifications c
                    JOIN transaction_lines l ON l.id=c.transaction_line_id
                    JOIN transaction_headers h ON h.id=l.header_id
                    WHERE l.cost_code_id IS NULL AND l.project_id=? AND h.party_id=? AND l.account_id=?
                      AND c.proposed_cost_code_id=?""",
                    (proj_id, party_id, acc_id, code_id)).fetchall():
                do_apply(cid, lid, code_id, "accept", "rule-group approval", None)

    # ---- Line Review (per-line) ----
    if "Line Review" in wb.sheetnames:
        ls = list(wb["Line Review"].iter_rows(values_only=True))
        lhdr = [str(c).strip() if c else "" for c in ls[0]]
        li = {h: k for k, h in enumerate(lhdr)}
        seen = set()
        for row in ls[1:]:
            if not row or not str(row[li["classification_id"]] or "").isdigit():
                continue
            cid = int(row[li["classification_id"]])
            decision = str(row[li["decision (accept/replace)"]] or "").strip().lower()
            if not decision:
                continue
            if cid in seen:
                res["rejected"] += 1; res["errors"].append(f"cls {cid}: duplicate decision"); continue
            seen.add(cid)
            crow = conn.execute("SELECT transaction_line_id, proposed_cost_code_id FROM classifications "
                                "WHERE id=?", (cid,)).fetchone()
            if not crow:
                res["rejected"] += 1; res["errors"].append(f"cls {cid}: unknown id"); continue
            line_id, proposed = crow
            if str(row[li["transaction_line_id"]]).strip() not in ("", str(line_id)):
                res["rejected"] += 1; res["errors"].append(f"cls {cid}: line id altered"); continue
            exported_hash = str(row[li["line_hash"]] or "").strip() or None
            comment = str(row[li["comment"]] or "")
            if decision == "accept":
                if proposed is None:
                    res["rejected"] += 1; res["errors"].append(f"cls {cid}: cannot accept abstain"); continue
                code_id = proposed
            elif decision == "replace":
                repl = str(row[li["replacement_code"]] or "").strip().lstrip("'")
                if repl not in code_by_str:
                    res["rejected"] += 1; res["errors"].append(f"cls {cid}: replacement '{repl}' invalid"); continue
                code_id = code_by_str[repl]
            else:
                res["rejected"] += 1; res["errors"].append(f"cls {cid}: bad decision"); continue
            do_apply(cid, line_id, code_id, decision, comment, exported_hash)

    # ---- conservative candidate promotion (>=2 consistent, conflict-free) ----
    for scope, codes in support.items():
        if scope in approved_scopes:
            continue                       # already an approved rule
        if len(codes) != 1:
            continue                       # conflicting decisions -> no promotion
        code_id, cnt = next(iter(codes.items()))
        if cnt < PROMOTE_MIN_SUPPORT:
            continue
        party_id, acc_id, proj_id = scope
        ex = conn.execute("""SELECT id, support_count FROM mapping_rules WHERE tenant_id=?
            AND IFNULL(party_id,-1)=IFNULL(?,-1) AND IFNULL(account_id,-1)=IFNULL(?,-1)
            AND IFNULL(project_id,-1)=IFNULL(?,-1) AND cost_code_id=?""",
            (tenant, party_id, acc_id, proj_id, code_id)).fetchone()
        if ex:
            conn.execute("UPDATE mapping_rules SET support_count=? WHERE id=?", (cnt, ex[0]))
        else:
            conn.execute("""INSERT INTO mapping_rules(tenant_id, party_id, account_id, project_id,
                cost_code_id, status, support_count, created_at)
                VALUES (?,?,?,?,?, 'candidate', ?, ?)""",
                (tenant, party_id, acc_id, proj_id, code_id, cnt, _now()))
            res["candidates"] += 1

    conn.execute("UPDATE review_batches SET status='applied' WHERE id=?", (batch_id,))
    if res["errors"]:
        now = _now()
        for e in res["errors"]:
            conn.execute("INSERT INTO data_quality_issues(severity, category, message, created_at) "
                         "VALUES ('warning','classification_review',?,?)", (e, now))
    conn.commit()
    conn.close()
    print(f"[review-apply] applied {res['applied']} · already {res['already']} · stale {res['stale']} · "
          f"rejected {res['rejected']} · rules approved {res['rules_approved']} · candidates {res['candidates']}")
    for e in res["errors"][:8]:
        print(f"           - {e}")
    return res


# --------------------------------------------------------------------------- #
# evaluation
# --------------------------------------------------------------------------- #

def evaluate_against_truth(db: str, truth: dict) -> dict:
    """truth: transaction_line_id -> true cost code. Reports per-method precision
    and the false-auto-apply rate (auto-applied lines whose code is wrong)."""
    conn = sqlite3.connect(db)
    rows = conn.execute("""SELECT c.method, cc.code, c.transaction_line_id
        FROM classifications c LEFT JOIN cost_codes cc ON cc.id=c.proposed_cost_code_id""").fetchall()
    conn.close()
    per = defaultdict(lambda: [0, 0])
    false_auto = 0
    for method, code, lid in rows:
        t = truth.get(lid)
        if t is None or code is None:
            continue
        per[method][1] += 1
        if code == t:
            per[method][0] += 1
        elif method in ("rule", "learned"):
            false_auto += 1
    out = {m: dict(correct=c, total=n, precision=(c / n if n else None))
           for m, (c, n) in per.items()}
    out["_false_auto_apply"] = false_auto
    return out
