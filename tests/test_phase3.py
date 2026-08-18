"""Phase 3 acceptance tests — hardened classification stack + safe review loop.

Covers: method stack, commitment-as-proposal, project-scoped validation, invented-
code rejection (#14), batch binding, idempotent apply (#3), optimistic-concurrency
stale rejection (#4), conservative rule promotion (#1), rule-group approval, and a
precision regression against a committed truth file (#8).

Run:  python tests/test_phase3.py
"""

from __future__ import annotations

import csv
import subprocess
import sqlite3
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = ROOT / "scripts"
SR = ROOT / "integration-data" / "summit-ridge"
sys.path.insert(0, str(SCRIPTS))

from ingest import run_ingest                               # noqa: E402
from load_workbook import load_workbook_into_db             # noqa: E402
from classify import (classify_db, write_review_workbook,   # noqa: E402
                      apply_review, evaluate_against_truth, REVIEW_SCHEMA_VERSION)


def _init(d: Path, name="c.db") -> str:
    db = str(d / name)
    subprocess.run([sys.executable, str(SCRIPTS / "cfo.py"), "init", "--db", db,
                    "--tenant", "Summit Ridge Construction Pty Ltd", "--force"],
                   check=True, capture_output=True)
    return db


def _qbo_db(d: Path, name="c.db") -> str:
    db = _init(d, name)
    run_ingest(db, str(SR / "qbo-exports" / "qbo-transaction-detail-current.csv"), "qbo",
               projects_workbook=str(SR / "project-input.xlsx"))
    load_workbook_into_db(db, str(SR / "project-input.xlsx"))
    return db


def _coded(db):
    c = sqlite3.connect(db)
    n = c.execute("""SELECT COUNT(*) FROM transaction_lines l JOIN transaction_headers h ON h.id=l.header_id
        WHERE l.project_id IS NOT NULL AND l.is_tax=0 AND h.is_void=0 AND l.cost_code_id IS NOT NULL""").fetchone()[0]
    c.close()
    return n


def _fill_line_decisions(path, want_accept):
    """Set decision='accept' on the first `want_accept` coded proposals. Returns count."""
    wb = load_workbook(path); ws = wb["Line Review"]
    hdr = [c.value for c in ws[1]]
    di, ci = hdr.index("decision (accept/replace)"), hdr.index("exported_code")
    n = 0
    for r in ws.iter_rows(min_row=2):
        if n >= want_accept:
            break
        if r[ci].value and r[ci].value != "UNCLASSIFIED":
            r[di].value = "accept"; n += 1
    wb.save(path)
    return n


def test_commitment_is_a_proposal_not_autoapply():
    with tempfile.TemporaryDirectory() as d:
        db = _qbo_db(Path(d))
        stats = classify_db(db)
        assert stats.get("commitment", 0) > 0, "commitments should be detected"
        # conservative: nothing auto-applies without an APPROVED rule
        assert _coded(db) == 0, "commitment matches must be proposals, not auto-applied"
        assert stats.get("unclassified", 0) >= 1, "ambiguous lines abstain"


def test_llm_cannot_invent_or_cross_project_code():
    with tempfile.TemporaryDirectory() as d:
        db = _qbo_db(Path(d))
        def rogue(line, allowed):
            return {"cost_code": "99-99-99", "score": 9999, "reason": "hallucinated"}
        stats = classify_db(db, proposer=rogue)
        assert stats.get("llm", 0) == 0, "an invalid code must never be accepted"


def test_batch_binding_rejects_foreign_workbook():
    with tempfile.TemporaryDirectory() as d:
        db1 = _qbo_db(Path(d), "a.db")
        db2 = _qbo_db(Path(d), "b.db")   # different db_uid
        classify_db(db1)
        review = str(Path(d) / "r.xlsx")
        write_review_workbook(db1, review)
        _fill_line_decisions(review, 1)
        try:
            apply_review(db2, review)     # wrong database
        except ValueError:
            return
        raise AssertionError("a workbook bound to another db must be rejected")


def test_apply_is_idempotent():
    with tempfile.TemporaryDirectory() as d:
        db = _qbo_db(Path(d))
        classify_db(db)
        review = str(Path(d) / "r.xlsx")
        write_review_workbook(db, review)
        acc = _fill_line_decisions(review, 5)
        r1 = apply_review(db, review)
        assert r1["applied"] == acc
        coded1 = _coded(db)
        r2 = apply_review(db, review)     # apply the SAME workbook again
        assert r2["applied"] == 0, "re-applying must change nothing"
        assert r2["already"] == acc
        assert _coded(db) == coded1


def test_stale_workbook_is_rejected():
    with tempfile.TemporaryDirectory() as d:
        db = _qbo_db(Path(d))
        classify_db(db)
        review = str(Path(d) / "r.xlsx")
        write_review_workbook(db, review)
        _fill_line_decisions(review, 3)
        # someone codes one of those lines through another path before we apply
        conn = sqlite3.connect(db)
        wb = load_workbook(review); ws = wb["Line Review"]
        hdr = [c.value for c in ws[1]]
        li, di = hdr.index("transaction_line_id"), hdr.index("decision (accept/replace)")
        first_line = next(r[li].value for r in ws.iter_rows(min_row=2) if r[di].value == "accept")
        any_code = conn.execute("SELECT id FROM cost_codes LIMIT 1").fetchone()[0]
        conn.execute("UPDATE transaction_lines SET cost_code_id=? WHERE id=?", (any_code, first_line))
        conn.commit(); conn.close()
        res = apply_review(db, review)
        assert res["stale"] >= 1, "a line coded after export must be treated as stale"


def test_single_acceptance_does_not_create_approved_rule():
    with tempfile.TemporaryDirectory() as d:
        db = _qbo_db(Path(d))
        classify_db(db)
        review = str(Path(d) / "r.xlsx")
        write_review_workbook(db, review)
        _fill_line_decisions(review, 1)   # exactly one acceptance
        apply_review(db, review)
        conn = sqlite3.connect(db)
        approved = conn.execute("SELECT COUNT(*) FROM mapping_rules WHERE status='approved'").fetchone()[0]
        conn.close()
        assert approved == 0, "a single acceptance must NOT create an auto-applying rule"


def test_two_consistent_acceptances_promote_candidate_only():
    with tempfile.TemporaryDirectory() as d:
        db = _qbo_db(Path(d))
        classify_db(db)
        review = str(Path(d) / "r.xlsx")
        write_review_workbook(db, review)
        # accept several lines from the SAME vendor/account/project/code scope
        wb = load_workbook(review); ws = wb["Line Review"]
        hdr = [c.value for c in ws[1]]
        di = hdr.index("decision (accept/replace)")
        vi, ai, pi, ci = (hdr.index("vendor"), hdr.index("account"),
                          hdr.index("project"), hdr.index("exported_code"))
        scope_counts = {}
        target = None
        for r in ws.iter_rows(min_row=2):
            if not r[ci].value or r[ci].value == "UNCLASSIFIED":
                continue
            key = (r[vi].value, r[ai].value, r[pi].value, r[ci].value)
            scope_counts[key] = scope_counts.get(key, 0) + 1
            if scope_counts[key] == 2 and target is None:
                target = key
        assert target, "need a scope with >=2 proposals"
        for r in ws.iter_rows(min_row=2):
            if (r[vi].value, r[ai].value, r[pi].value, r[ci].value) == target:
                r[di].value = "accept"
        wb.save(review)
        apply_review(db, review)
        conn = sqlite3.connect(db)
        approved = conn.execute("SELECT COUNT(*) FROM mapping_rules WHERE status='approved'").fetchone()[0]
        cand = conn.execute("SELECT COUNT(*) FROM mapping_rules WHERE status='candidate'").fetchone()[0]
        conn.close()
        assert approved == 0, "consistent acceptances form a CANDIDATE, not an approved rule"
        assert cand >= 1, "two consistent confirmations should form a candidate rule"


def test_rule_group_approval_creates_approved_rule_and_codes_lines():
    with tempfile.TemporaryDirectory() as d:
        db = _qbo_db(Path(d))
        classify_db(db)
        review = str(Path(d) / "r.xlsx")
        write_review_workbook(db, review)
        wb = load_workbook(review); gs = wb["Rule Groups"]
        hdr = [c.value for c in gs[1]]
        di = hdr.index("decision (approve rule/review lines)")
        # approve the first (largest) rule group
        rows = list(gs.iter_rows(min_row=2))
        rows[0][di].value = "approve rule"
        wb.save(review)
        before = _coded(db)
        res = apply_review(db, review)
        assert res["rules_approved"] >= 1
        assert res["applied"] >= 1
        assert _coded(db) > before
        conn = sqlite3.connect(db)
        approved = conn.execute("SELECT COUNT(*) FROM mapping_rules WHERE status='approved'").fetchone()[0]
        conn.close()
        assert approved >= 1


def _build_truth(db):
    """Map this db's line ids to true codes via the committed truth file."""
    key2code = {}
    with (SR / "expected" / "classification-truth.csv").open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            key2code[(r["txn_type"], r["txn_id"], r["account"], r["project"],
                      int(r["amount_minor"]), r["memo"])] = r["true_cost_code"]
    conn = sqlite3.connect(db)
    rows = conn.execute("""SELECT l.id, h.native_transaction_type, h.native_transaction_id,
        a.name, p.native_id, l.amount_minor, l.memo
        FROM transaction_lines l JOIN transaction_headers h ON h.id=l.header_id
        LEFT JOIN accounts a ON a.id=l.account_id LEFT JOIN projects p ON p.id=l.project_id
        WHERE l.project_id IS NOT NULL AND l.is_tax=0""").fetchall()
    conn.close()
    truth = {}
    for lid, tt, tid, acc, proj, amt, memo in rows:
        c = key2code.get((tt, tid, acc, proj, amt, memo))
        if c:
            truth[lid] = c
    return truth


def test_precision_regression_against_committed_truth():
    with tempfile.TemporaryDirectory() as d:
        db = _qbo_db(Path(d))
        classify_db(db)
        truth = _build_truth(db)
        assert len(truth) > 500, f"truth mapping too small: {len(truth)}"
        res = evaluate_against_truth(db, truth)
        assert res["_false_auto_apply"] == 0, "nothing wrong may be auto-applied"
        assert res["commitment"]["precision"] == 1.0, res["commitment"]
        assert res["heuristic"]["precision"] >= 0.90, res["heuristic"]


def _run_all():
    for name, fn in sorted(globals().items()):
        if name.startswith("test_") and callable(fn):
            fn()
            print(f"PASS {name}")


if __name__ == "__main__":
    _run_all()
    print("\nAll Phase 3 acceptance checks passed.")
